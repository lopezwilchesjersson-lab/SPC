"""
Pulso de Calidad SPC

Plataforma visual para Control Estadístico de Procesos.
Creado por Jerson Andrés López Wilches.

Instalar dependencias:
    python app.py --install

Probar funciones internas:
    python app.py --self-test

Ejecutar:
    python -m streamlit run app.py
"""

from __future__ import annotations

import sys
import time
import subprocess
import importlib.util
from io import StringIO, BytesIO
from typing import Optional
import tempfile
import os

PAQUETES_APP = [
    "streamlit",
    "pandas",
    "numpy",
    "plotly",
    "scipy",
    "statsmodels",
    "openpyxl",
]

PAQUETES_BASE = ["pandas", "numpy", "scipy", "statsmodels"]


def existe_paquete(nombre: str) -> bool:
    return importlib.util.find_spec(nombre) is not None


def paquetes_faltantes(paquetes: list[str]) -> list[str]:
    return [p for p in paquetes if not existe_paquete(p)]


def instalar_paquetes(paquetes: list[str]) -> bool:
    if not paquetes:
        print("Todas las librerías ya están instaladas.")
        return True
    comando = [sys.executable, "-m", "pip", "install", *paquetes]
    print("Instalando librerías:", ", ".join(paquetes))
    try:
        subprocess.check_call(comando)
        print("Instalación completada.")
        print("Ejecuta: python -m streamlit run app.py")
        return True
    except Exception as error:
        print(f"No se pudo instalar: {error}")
        print("Ejecuta manualmente:")
        print("python -m pip install streamlit pandas numpy scipy statsmodels plotly openpyxl")
        return False


def mensaje_dependencias(faltantes: list[str]) -> str:
    lineas = [
        "Faltan librerías para ejecutar la aplicación.",
        "",
        "Librerías faltantes:",
        *[f"- {p}" for p in faltantes],
        "",
        "Solución:",
        "python -m pip install streamlit pandas numpy scipy statsmodels plotly openpyxl",
        "",
        "Luego ejecuta:",
        "python -m streamlit run app.py",
    ]
    return "\n".join(lineas)


if "--install" in sys.argv:
    instalar_paquetes(paquetes_faltantes(PAQUETES_APP))
    raise SystemExit(0)

FALTAN_BASE = paquetes_faltantes(PAQUETES_BASE)
FALTAN_APP = paquetes_faltantes(PAQUETES_APP)

if FALTAN_BASE:
    print(mensaje_dependencias(FALTAN_BASE))
else:
    import pandas as pd
    import numpy as np
    from scipy import stats
    from statsmodels.stats.stattools import durbin_watson


# ============================================================
# UTILIDADES GENERALES
# ============================================================

def verificar_dependencias_base() -> None:
    if FALTAN_BASE:
        raise RuntimeError(mensaje_dependencias(FALTAN_BASE))


def limpiar_nombre_columna(nombre) -> str:
    nombre = str(nombre).strip()
    reemplazos = {
        "LSL": "LIE",
        "USL": "LSE",
        "xbar": "Xbarra",
        "x-bar": "Xbarra",
    }
    return reemplazos.get(nombre, nombre)


def normalizar_nombres_columnas(df):
    verificar_dependencias_base()
    df = df.copy()
    nuevas = []
    usados = {}
    for col in df.columns:
        nombre = limpiar_nombre_columna(col)
        if nombre in usados:
            usados[nombre] += 1
            nombre = f"{nombre}_{usados[nombre]}"
        else:
            usados[nombre] = 0
        nuevas.append(nombre)
    df.columns = nuevas
    return df


def dataframe_valido(df) -> bool:
    verificar_dependencias_base()
    return df is not None and isinstance(df, pd.DataFrame) and not df.empty


def obtener_columna(df, columna):
    verificar_dependencias_base()
    datos = df.loc[:, columna]
    if isinstance(datos, pd.DataFrame):
        datos = datos.iloc[:, 0]
    return datos


def convertir_a_numerica(serie):
    verificar_dependencias_base()
    if isinstance(serie, pd.DataFrame):
        serie = serie.iloc[:, 0]
    return pd.to_numeric(serie, errors="coerce").dropna()


def columnas_numericas(df) -> list[str]:
    verificar_dependencias_base()
    if not dataframe_valido(df):
        return []
    cols = []
    for col in df.columns:
        if len(convertir_a_numerica(obtener_columna(df, col))) > 0:
            cols.append(col)
    return cols


def fmt(valor, decimales: Optional[int] = None):
    if isinstance(valor, (float, int, np.integer, np.floating)) and not pd.isna(valor):
        valor = float(valor)
        if valor == 0:
            return "0"
        if abs(valor) >= 100000:
            return f"{valor:,.0f}"
        if abs(valor) >= 1000:
            return f"{valor:,.1f}".rstrip("0").rstrip(".")
        if abs(valor) >= 100:
            return f"{valor:,.2f}".rstrip("0").rstrip(".")
        if abs(valor) >= 10:
            return f"{valor:,.3f}".rstrip("0").rstrip(".")
        if abs(valor) >= 1:
            return f"{valor:,.4f}".rstrip("0").rstrip(".")
        if abs(valor) >= 0.01:
            return f"{valor:,.5f}".rstrip("0").rstrip(".")
        return f"{valor:,.8f}".rstrip("0").rstrip(".")
    return str(valor)


def valor_limpio(valor):
    if isinstance(valor, (float, int, np.integer, np.floating)) and not pd.isna(valor):
        texto = fmt(valor)
        try:
            return float(texto.replace(",", ""))
        except Exception:
            return texto
    return valor


def redondear_dict(datos: dict, decimales: Optional[int] = None) -> dict:
    salida = {}
    for k, v in datos.items():
        salida[k] = valor_limpio(v)
    return salida


def guardar_estado_si_no_existe(st, nombre, valor):
    if nombre not in st.session_state:
        st.session_state[nombre] = valor


def _actualizar_estado_persistente(nombre):
    import streamlit as st
    clave_widget = f"_widget_{nombre}"
    if clave_widget in st.session_state:
        st.session_state[nombre] = st.session_state[clave_widget]


def numero_persistente(st, etiqueta, nombre, valor_defecto, **kwargs):
    guardar_estado_si_no_existe(st, nombre, valor_defecto)
    clave_widget = f"_widget_{nombre}"
    if clave_widget not in st.session_state:
        st.session_state[clave_widget] = st.session_state[nombre]
    return st.number_input(
        etiqueta,
        key=clave_widget,
        on_change=_actualizar_estado_persistente,
        args=(nombre,),
        **kwargs,
    )


def checkbox_persistente(st, etiqueta, nombre, valor_defecto=False, **kwargs):
    guardar_estado_si_no_existe(st, nombre, valor_defecto)
    clave_widget = f"_widget_{nombre}"
    if clave_widget not in st.session_state:
        st.session_state[clave_widget] = st.session_state[nombre]
    return st.checkbox(
        etiqueta,
        key=clave_widget,
        on_change=_actualizar_estado_persistente,
        args=(nombre,),
        **kwargs,
    )


def selectbox_persistente(st, etiqueta, opciones, nombre, valor_defecto=None, **kwargs):
    if valor_defecto is None:
        valor_defecto = opciones[0]
    if valor_defecto not in opciones:
        valor_defecto = opciones[0]
    guardar_estado_si_no_existe(st, nombre, valor_defecto)
    if st.session_state[nombre] not in opciones:
        st.session_state[nombre] = opciones[0]
    clave_widget = f"_widget_{nombre}"
    if clave_widget not in st.session_state or st.session_state[clave_widget] not in opciones:
        st.session_state[clave_widget] = st.session_state[nombre]
    indice = opciones.index(st.session_state[clave_widget])
    return st.selectbox(
        etiqueta,
        opciones,
        index=indice,
        key=clave_widget,
        on_change=_actualizar_estado_persistente,
        args=(nombre,),
        **kwargs,
    )


def radio_persistente(st, etiqueta, opciones, nombre, valor_defecto=None, **kwargs):
    if valor_defecto is None:
        valor_defecto = opciones[0]
    guardar_estado_si_no_existe(st, nombre, valor_defecto)
    if st.session_state[nombre] not in opciones:
        st.session_state[nombre] = opciones[0]
    clave_widget = f"_widget_{nombre}"
    if clave_widget not in st.session_state or st.session_state[clave_widget] not in opciones:
        st.session_state[clave_widget] = st.session_state[nombre]
    indice = opciones.index(st.session_state[clave_widget])
    return st.radio(
        etiqueta,
        opciones,
        index=indice,
        key=clave_widget,
        on_change=_actualizar_estado_persistente,
        args=(nombre,),
        **kwargs,
    )


def cargar_datos_desde_archivo(archivo):
    verificar_dependencias_base()
    nombre = archivo.name.lower()
    if nombre.endswith(".csv"):
        df = pd.read_csv(archivo)
    elif nombre.endswith(".xlsx"):
        df = pd.read_excel(archivo)
    else:
        raise ValueError("Formato no soportado. Usa CSV o XLSX.")
    return normalizar_nombres_columnas(df)


def convertir_ancho_a_largo(df, columna_subgrupo):
    verificar_dependencias_base()
    if columna_subgrupo not in df.columns:
        return df
    df_largo = df.melt(id_vars=columna_subgrupo, var_name="medicion_id", value_name="medicion")
    df_largo = df_largo.rename(columns={columna_subgrupo: "subgrupo"})
    df_largo["medicion"] = pd.to_numeric(df_largo["medicion"], errors="coerce")
    df_largo = df_largo.dropna(subset=["medicion"])
    return normalizar_nombres_columnas(df_largo)


# ============================================================
# CÁLCULOS ESTADÍSTICOS
# ============================================================

def resumen_descriptivo(serie):
    verificar_dependencias_base()
    s = convertir_a_numerica(serie)
    if len(s) == 0:
        return None
    return {
        "n": int(len(s)),
        "Media": float(s.mean()),
        "Mediana": float(s.median()),
        "Desv. estándar": float(s.std(ddof=1)) if len(s) > 1 else np.nan,
        "Mínimo": float(s.min()),
        "Máximo": float(s.max()),
        "Rango": float(s.max() - s.min()),
        "Q1": float(s.quantile(0.25)),
        "Q3": float(s.quantile(0.75)),
    }


def detectar_atipicos_iqr(serie):
    verificar_dependencias_base()
    s = convertir_a_numerica(serie)
    if len(s) == 0:
        return np.nan, np.nan, s
    q1 = s.quantile(0.25)
    q3 = s.quantile(0.75)
    iqr = q3 - q1
    li = q1 - 1.5 * iqr
    ls = q3 + 1.5 * iqr
    return float(li), float(ls), s[(s < li) | (s > ls)]


def evaluar_normalidad(serie):
    verificar_dependencias_base()
    s = convertir_a_numerica(serie)
    if len(s) < 3:
        return pd.DataFrame([{"Prueba": "Normalidad", "p-valor": np.nan, "Resultado": "Datos insuficientes", "Interpretación": "Se necesitan al menos 3 datos."}])
    pruebas = [("Shapiro-Wilk", stats.shapiro)]
    if len(s) >= 8:
        pruebas.append(("D'Agostino-Pearson", stats.normaltest))
    pruebas.append(("Jarque-Bera", stats.jarque_bera))
    filas = []
    for nombre, funcion in pruebas:
        try:
            est, p = funcion(s)
            resultado = "Cumple" if p >= 0.05 else "No cumple"
            filas.append({
                "Prueba": nombre,
                "Estadístico": round(float(est), 3),
                "p-valor": round(float(p), 3),
                "Resultado": resultado,
                "Interpretación": "Compatible con normalidad." if resultado == "Cumple" else "No se ajusta claramente a normalidad.",
            })
        except Exception as e:
            filas.append({"Prueba": nombre, "Estadístico": np.nan, "p-valor": np.nan, "Resultado": "No evaluado", "Interpretación": str(e)})
    return pd.DataFrame(filas)


def diagnostico_normalidad(tabla):
    if tabla is None or tabla.empty:
        return "No evaluado", "No hay resultados de normalidad."
    pvals = pd.to_numeric(tabla["p-valor"], errors="coerce").dropna()
    if len(pvals) == 0:
        return "No evaluado", "No hay p-valores suficientes."
    if (pvals >= 0.05).mean() >= 0.5:
        return "Cumple", "La mayoría de pruebas no rechaza normalidad con α = 0.05."
    return "No cumple", "La mayoría de pruebas rechaza normalidad con α = 0.05."


def evaluar_independencia(serie):
    verificar_dependencias_base()
    s = convertir_a_numerica(serie).reset_index(drop=True)
    if len(s) < 4:
        return {"Durbin-Watson": np.nan, "Autocorr. lag 1": np.nan, "Resultado": "Datos insuficientes", "Interpretación": "Se necesitan al menos 4 datos."}
    try:
        dw = float(durbin_watson(s))
        ac = float(s.autocorr(lag=1))
    except Exception as e:
        return {"Durbin-Watson": np.nan, "Autocorr. lag 1": np.nan, "Resultado": "No evaluado", "Interpretación": str(e)}
    cumple = 1.5 <= dw <= 2.5 and (not pd.isna(ac)) and abs(ac) <= 0.30
    return {
        "Durbin-Watson": round(dw, 3),
        "Autocorr. lag 1": round(ac, 3) if not pd.isna(ac) else np.nan,
        "Resultado": "Cumple" if cumple else "No cumple",
        "Interpretación": "No se detecta autocorrelación fuerte." if cumple else "Existe posible dependencia temporal, racha, tendencia o ciclo.",
    }


def evaluar_homocedasticidad(df, columna_valor, columna_grupo):
    verificar_dependencias_base()
    if not dataframe_valido(df) or columna_valor not in df.columns or columna_grupo not in df.columns:
        return pd.DataFrame([{"Prueba": "Homocedasticidad", "Resultado": "Columnas no válidas"}])
    grupos = []
    for _, g in df.groupby(columna_grupo):
        vals = convertir_a_numerica(obtener_columna(g, columna_valor))
        if len(vals) >= 2:
            grupos.append(vals)
    if len(grupos) < 2:
        return pd.DataFrame([{"Prueba": "Homocedasticidad", "Resultado": "Datos insuficientes", "Interpretación": "Se requieren al menos 2 grupos con 2 datos."}])
    filas = []
    for nombre, funcion in [("Levene", stats.levene), ("Bartlett", stats.bartlett)]:
        try:
            est, p = funcion(*grupos)
            resultado = "Cumple" if p >= 0.05 else "No cumple"
            filas.append({
                "Prueba": nombre,
                "Estadístico": round(float(est), 3),
                "p-valor": round(float(p), 3),
                "Resultado": resultado,
                "Interpretación": "Varianzas similares." if resultado == "Cumple" else "Varianzas diferentes.",
            })
        except Exception as e:
            filas.append({"Prueba": nombre, "Resultado": "No evaluado", "Interpretación": str(e)})
    return pd.DataFrame(filas)


# ============================================================
# CAPACIDAD
# ============================================================

def clasificar_capacidad(cpk: float) -> str:
    if cpk >= 1.67:
        return "Excelente"
    if cpk >= 1.33:
        return "Capaz"
    if cpk >= 1.00:
        return "Marginal"
    return "No capaz"


def calcular_capacidad(serie, lie, lse, vn=None, sigma_historica=None, media_conocida=None):
    verificar_dependencias_base()
    s = convertir_a_numerica(serie)
    if len(s) < 2 or lie >= lse:
        return None
    media_muestral = float(s.mean())
    media = float(media_conocida) if media_conocida is not None else media_muestral
    sigma_m = float(s.std(ddof=1))
    sigma = float(sigma_historica) if sigma_historica is not None and sigma_historica > 0 else sigma_m
    if sigma <= 0 or pd.isna(sigma):
        return None
    if vn is None:
        vn = (lie + lse) / 2
    cp = (lse - lie) / (6 * sigma)
    cpu = (lse - media) / (3 * sigma)
    cpl = (media - lie) / (3 * sigma)
    cpk = min(cpu, cpl)
    pp = (lse - lie) / (6 * sigma_m)
    ppu = (lse - media) / (3 * sigma_m)
    ppl = (media - lie) / (3 * sigma_m)
    ppk = min(ppu, ppl)
    cpm = (lse - lie) / (6 * np.sqrt(sigma**2 + (media - vn) ** 2))
    obs_lie = int((s < lie).sum())
    obs_lse = int((s > lse).sum())
    obs_total = obs_lie + obs_lse
    p_lie = float(stats.norm.cdf(lie, loc=media, scale=sigma))
    p_lse = float(1 - stats.norm.cdf(lse, loc=media, scale=sigma))
    p_total = p_lie + p_lse
    return {
        "n": int(len(s)),
        "VN": float(vn),
        "LIE": float(lie),
        "LSE": float(lse),
        "Media usada": media,
        "Media muestral": media_muestral,
        "Media conocida usada": "Sí" if media_conocida is not None else "No",
        "Sigma usada": sigma,
        "Sigma muestral": sigma_m,
        "Sigma histórica usada": "Sí" if sigma_historica is not None and sigma_historica > 0 else "No",
        "Cp": float(cp),
        "CPU": float(cpu),
        "CPL": float(cpl),
        "Cpk": float(cpk),
        "Pp": float(pp),
        "PPU": float(ppu),
        "PPL": float(ppl),
        "Ppk": float(ppk),
        "Cpm": float(cpm),
        "Z LIE": float((media - lie) / sigma),
        "Z LSE": float((lse - media) / sigma),
        "Producto no conforme observado": int(obs_total),
        "Fuera por LIE observado": int(obs_lie),
        "Fuera por LSE observado": int(obs_lse),
        "% PNC observado": float(obs_total / len(s) * 100),
        "% PNC estimado LIE": p_lie * 100,
        "% PNC estimado LSE": p_lse * 100,
        "% PNC estimado total": p_total * 100,
        "PPM estimado": p_total * 1_000_000,
        "Estado": clasificar_capacidad(cpk),
        "Centrado": "Centrado" if abs(media - vn) <= 0.25 * sigma else "Descentrado",
        "Riesgo principal": "LSE" if cpu < cpl else "LIE" if cpl < cpu else "Equilibrado",
    }


def diagnostico_capacidad(res):
    if res is None:
        return ["No se pudo calcular capacidad. Revisa que existan datos numéricos, que LIE sea menor que LSE y que la desviación estándar sea mayor que cero."]

    mensajes = []
    cpk = res["Cpk"]
    cp = res["Cp"]
    cpl = res["CPL"]
    cpu = res["CPU"]
    pnc_total = res["% PNC estimado total"]
    pnc_obs = res["% PNC observado"]

    mensajes.append(f"Conclusión general: el proceso se clasifica como {res['Estado']}. Esta clasificación se basa principalmente en Cpk, porque Cpk considera al mismo tiempo la variabilidad y el centrado frente a los límites de especificación.")

    if cp < 1:
        mensajes.append("Cp es menor que 1. Esto indica que la variabilidad natural del proceso es más ancha que la tolerancia disponible entre LIE y LSE. Aunque el proceso estuviera centrado, seguiría teniendo riesgo alto de generar producto fuera de especificación.")
    elif cp < 1.33:
        mensajes.append("Cp está entre 1 y 1.33. La variabilidad cabe dentro de la tolerancia, pero el margen es limitado. El proceso requiere control cuidadoso y reducción de variación para trabajar con seguridad.")
    else:
        mensajes.append("Cp es igual o mayor que 1.33. La variabilidad potencial del proceso es adecuada frente a la tolerancia, siempre que el proceso esté estable y bien centrado.")

    if cpk < cp * 0.85:
        mensajes.append("Cpk es bastante menor que Cp. Esto evidencia descentrado: el problema no es solo la variabilidad, sino también la ubicación de la media dentro de los límites de especificación.")
    else:
        mensajes.append("Cpk es cercano a Cp. Esto indica que el proceso está relativamente centrado respecto a sus límites de especificación.")

    if cpu < cpl:
        mensajes.append("CPU es menor que CPL. El lado crítico es el límite superior de especificación LSE. La media está más cerca del límite superior, por eso el mayor riesgo es producir valores altos fuera de especificación.")
    elif cpl < cpu:
        mensajes.append("CPL es menor que CPU. El lado crítico es el límite inferior de especificación LIE. La media está más cerca del límite inferior, por eso el mayor riesgo es producir valores bajos fuera de especificación.")
    else:
        mensajes.append("CPU y CPL son similares. El riesgo está equilibrado entre el límite inferior y el límite superior.")

    if res["Centrado"] == "Descentrado":
        mensajes.append("El proceso está descentrado respecto al valor nominal. Para mejorar Cpk no basta con reducir variación; también se debe ajustar la media del proceso hacia el objetivo o hacia el centro de especificación.")
    else:
        mensajes.append("El proceso aparece centrado respecto al valor nominal. En este caso, la mejora principal debe enfocarse en reducir variabilidad si Cp o Cpk son bajos.")

    if pnc_total > 5:
        mensajes.append(f"El producto no conforme estimado es alto: {fmt(pnc_total)}%. Este nivel indica pérdidas relevantes, reproceso o incumplimiento frecuente de especificaciones.")
    elif pnc_total > 1:
        mensajes.append(f"El producto no conforme estimado es moderado: {fmt(pnc_total)}%. Conviene mejorar el proceso antes de usarlo como proceso estable de producción.")
    else:
        mensajes.append(f"El producto no conforme estimado es bajo: {fmt(pnc_total)}%. Aun así, debe verificarse estabilidad con cartas de control antes de concluir capacidad final.")

    if abs(pnc_obs - pnc_total) <= max(1, pnc_total * 0.25):
        mensajes.append("El PNC observado y el PNC estimado son cercanos. Esto sugiere coherencia entre los datos observados y el modelo normal usado en capacidad.")
    else:
        mensajes.append("El PNC observado y el PNC estimado difieren de forma importante. Revisa normalidad, datos atípicos, estabilidad del proceso o si la distribución real no es normal.")

    if res["Producto no conforme observado"] > 0:
        mensajes.append(f"En los datos se observaron {res['Producto no conforme observado']} unidades fuera de especificación: {res['Fuera por LIE observado']} por debajo de LIE y {res['Fuera por LSE observado']} por encima de LSE.")
    else:
        mensajes.append("En la muestra no se observaron unidades fuera de especificación. Esto no garantiza capacidad si Cp o Cpk son bajos; solo indica que la muestra analizada no presentó fallos observados.")

    mensajes.append("Recomendación técnica: primero confirma estabilidad con la carta de control correspondiente. Después ajusta el centrado de la media y reduce la variabilidad. La capacidad solo debe interpretarse como válida cuando el proceso está bajo control estadístico.")
    return mensajes


def calcular_sigma_objetivo(lie, lse, cp_objetivo):
    if cp_objetivo <= 0:
        return np.nan
    return float((lse - lie) / (6 * cp_objetivo))


def media_maxima_para_pnc_lse(lse, sigma, pnc_lse):
    z = stats.norm.ppf(1 - pnc_lse)
    return float(lse - z * sigma)


def potencia_xbarra(media_actual, media_cambio, sigma, n, z=3):
    if sigma <= 0 or n <= 0:
        return None
    diferencia = abs(media_cambio - media_actual)
    d = diferencia / sigma
    d_raiz_n = d * np.sqrt(n)
    z_menos = z - d_raiz_n
    beta = float(stats.norm.cdf(z_menos))
    potencia = float(1 - beta)
    arl = np.inf if potencia <= 0 else 1 / potencia
    return {
        "Media actual": float(media_actual),
        "Media a detectar": float(media_cambio),
        "Sigma usada": float(sigma),
        "Diferencia absoluta": float(diferencia),
        "d": float(d),
        "d√n": float(d_raiz_n),
        "Z crítico": float(z),
        "Z crítico - d√n": float(z_menos),
        "β": float(beta),
        "% β": float(beta * 100),
        "Potencia = 1 - β": potencia,
        "% Potencia": potencia * 100,
        "ARL1": arl,
    }


def n_para_potencia(media_actual, media_cambio, sigma, potencia_objetivo, z=3):
    if sigma <= 0 or potencia_objetivo <= 0 or potencia_objetivo >= 1 or media_actual == media_cambio:
        return None
    d = abs(media_cambio - media_actual) / sigma
    z_beta = stats.norm.ppf(1 - potencia_objetivo)
    return int(np.ceil(((z - z_beta) / d) ** 2))



def _proporcion_desde_unidad(valor, unidad):
    valor = float(valor)
    if unidad == "Porcentaje defectuoso (%)":
        return valor / 100
    if unidad == "Defectivos por millón (DPM)":
        return valor / 1_000_000
    return valor


def _valor_a_unidad(p, unidad):
    p = float(p)
    if unidad == "Porcentaje defectuoso (%)":
        return p * 100
    if unidad == "Defectivos por millón (DPM)":
        return p * 1_000_000
    return p


def prob_aceptacion_atributos(n, c, p, distribucion="Binomial", lote=None):
    """Probabilidad de aceptación del lote: P(X <= c)."""
    n = int(n)
    c = int(c)
    p = float(min(max(p, 0.0), 1.0))
    if n <= 0 or c < 0:
        return np.nan
    if c >= n:
        return 1.0
    if distribucion == "Hipergeométrica" and lote is not None and lote > 0:
        N = int(lote)
        n = min(n, N)
        D = int(round(p * N))
        D = max(0, min(D, N))
        return float(stats.hypergeom.cdf(c, N, D, n))
    return float(stats.binom.cdf(c, n, p))


def _c_min_para_pa(n, p, pa_min, distribucion="Binomial", lote=None):
    n = int(n)
    if distribucion == "Hipergeométrica" and lote is not None and lote > 0:
        N = int(lote)
        D = int(round(p * N))
        D = max(0, min(D, N))
        for c in range(0, n + 1):
            if stats.hypergeom.cdf(c, N, D, n) >= pa_min:
                return c
        return n
    c = int(np.ceil(stats.binom.ppf(pa_min, n, p)))
    c = max(0, min(c, n))
    while c > 0 and stats.binom.cdf(c - 1, n, p) >= pa_min:
        c -= 1
    while c < n and stats.binom.cdf(c, n, p) < pa_min:
        c += 1
    return c


def _c_max_para_pa(n, p, pa_max, distribucion="Binomial", lote=None):
    n = int(n)
    if distribucion == "Hipergeométrica" and lote is not None and lote > 0:
        N = int(lote)
        D = int(round(p * N))
        D = max(0, min(D, N))
        mejor = -1
        for c in range(0, n + 1):
            if stats.hypergeom.cdf(c, N, D, n) <= pa_max:
                mejor = c
            else:
                break
        return mejor
    c = int(np.floor(stats.binom.ppf(pa_max, n, p)))
    c = max(-1, min(c, n))
    while c >= 0 and stats.binom.cdf(c, n, p) > pa_max:
        c -= 1
    while c + 1 <= n and stats.binom.cdf(c + 1, n, p) <= pa_max:
        c += 1
    return c


def validar_plan_atributos(n, c, aql, rql, alfa=0.05, beta=0.10, lote=None, distribucion="Binomial"):
    pa_aql = prob_aceptacion_atributos(n, c, aql, distribucion, lote)
    pa_rql = prob_aceptacion_atributos(n, c, rql, distribucion, lote)
    return {
        "Muestra n": int(n),
        "Aceptación c": int(c),
        "Rechazo r": int(c) + 1,
        "AQL": float(aql),
        "RQL": float(rql),
        "Alfa objetivo": float(alfa),
        "Beta objetivo": float(beta),
        "Pa en AQL": pa_aql,
        "Pa en RQL": pa_rql,
        "Riesgo productor real": 1 - pa_aql,
        "Riesgo consumidor real": pa_rql,
        "Distribución": distribucion,
        "Lote": int(lote) if lote is not None and lote > 0 else "No indicado",
        "Cumple": "Sí" if (pa_aql >= 1 - alfa and pa_rql <= beta) else "No",
        "Origen": "Plan validado con los n y c ingresados.",
    }


def disenar_plan_atributos_r_minitab(aql, rql, alfa=0.05, beta=0.10, lote=None, distribucion="Binomial", n_max=None):
    """Diseña el menor plan (n, c) que cumple las condiciones de calidad y riesgo definidas."""
    aql = float(min(max(aql, 0.0000001), 0.999999))
    rql = float(min(max(rql, aql + 0.0000001), 0.999999))
    alfa = float(min(max(alfa, 0.000001), 0.499999))
    beta = float(min(max(beta, 0.000001), 0.499999))
    if n_max is None:
        n_max = int(lote) if lote is not None and lote > 0 else 50000
    n_max = int(max(1, n_max))
    if distribucion == "Hipergeométrica" and lote is not None and lote > 0:
        n_max = min(n_max, int(lote))

    pa_aql_min = 1 - alfa
    pa_rql_max = beta
    for n in range(1, n_max + 1):
        c_min = _c_min_para_pa(n, aql, pa_aql_min, distribucion, lote)
        c_max = _c_max_para_pa(n, rql, pa_rql_max, distribucion, lote)
        if c_min <= c_max and c_min >= 0:
            c = int(c_min)
            plan = validar_plan_atributos(n, c, aql, rql, alfa, beta, lote, distribucion)
            plan["Cumple"] = "Sí"
            plan["Origen"] = "Menor n encontrado que cumple el punto del productor y el punto del consumidor."
            return plan

    n = n_max
    c = _c_min_para_pa(n, aql, pa_aql_min, distribucion, lote)
    plan = validar_plan_atributos(n, c, aql, rql, alfa, beta, lote, distribucion)
    plan["Cumple"] = "No"
    plan["Origen"] = "No se encontró un plan que cumpla ambos riesgos dentro del límite de búsqueda."
    return plan


def tabla_oc_atributos(plan, puntos=140):
    rql = float(plan["RQL"])
    max_p = min(max(rql * 2.2, 0.10), 0.50)
    ps = np.linspace(0, max_p, puntos)
    lote = plan["Lote"] if plan["Lote"] != "No indicado" else None
    return pd.DataFrame({
        "Fracción defectuosa": ps,
        "Probabilidad de aceptación": [prob_aceptacion_atributos(plan["Muestra n"], plan["Aceptación c"], p, plan["Distribución"], lote) for p in ps]
    })


def texto_plan_muestreo_minitab(plan, producto="unidades"):
    lote = plan.get("Lote", "No indicado")
    n = int(plan["Muestra n"])
    c = int(plan["Aceptación c"])
    r = int(plan["Rechazo r"])
    if lote == "No indicado":
        inicio = f"Se debe tomar una muestra aleatoria de {n} {producto}."
    else:
        inicio = f"Para un lote de {lote} {producto}, se debe tomar una muestra aleatoria de {n} {producto}."
    return (
        f"{inicio} Se acepta el lote si se encuentran {c} o menos unidades no conformes. "
        f"Se rechaza el lote si se encuentran {r} o más unidades no conformes. "
        f"Con este plan, Pa en AQL es {fmt(plan['Pa en AQL'] * 100)}% "
        f"y Pa en RQL/LTPD es {fmt(plan['Pa en RQL'] * 100)}%."
    )

# ============================================================
# CARTAS DE CONTROL
# ============================================================

def constantes_xbar_r(n):
    tabla = {
        2: {"A2": 1.880, "D3": 0.000, "D4": 3.267, "d2": 1.128},
        3: {"A2": 1.023, "D3": 0.000, "D4": 2.574, "d2": 1.693},
        4: {"A2": 0.729, "D3": 0.000, "D4": 2.282, "d2": 2.059},
        5: {"A2": 0.577, "D3": 0.000, "D4": 2.114, "d2": 2.326},
        6: {"A2": 0.483, "D3": 0.000, "D4": 2.004, "d2": 2.534},
        7: {"A2": 0.419, "D3": 0.076, "D4": 1.924, "d2": 2.704},
        8: {"A2": 0.373, "D3": 0.136, "D4": 1.864, "d2": 2.847},
        9: {"A2": 0.337, "D3": 0.184, "D4": 1.816, "d2": 2.970},
        10: {"A2": 0.308, "D3": 0.223, "D4": 1.777, "d2": 3.078},
    }
    return tabla.get(int(n))


def calcular_xbar_r(df, col_valor, col_subgrupo):
    if not dataframe_valido(df) or col_valor not in df.columns or col_subgrupo not in df.columns:
        return None
    temp = pd.DataFrame({
        "valor": pd.to_numeric(obtener_columna(df, col_valor), errors="coerce"),
        "subgrupo": obtener_columna(df, col_subgrupo),
    }).dropna()
    if temp.empty:
        return None
    resumen = temp.groupby("subgrupo")["valor"].agg(["mean", "max", "min", "count"])
    resumen["R"] = resumen["max"] - resumen["min"]
    resumen = resumen[resumen["count"] >= 2]
    if resumen.empty:
        return None
    n = int(round(float(resumen["count"].mean())))
    c = constantes_xbar_r(n)
    if not c:
        return None
    xbb = float(resumen["mean"].mean())
    rb = float(resumen["R"].mean())
    limites = {
        "Tamaño promedio de subgrupo": n,
        "X doble barra": xbb,
        "R barra": rb,
        "A2": c["A2"],
        "D3": c["D3"],
        "D4": c["D4"],
        "d2": c["d2"],
        "LC Xbarra": xbb,
        "LCS Xbarra": xbb + c["A2"] * rb,
        "LCI Xbarra": xbb - c["A2"] * rb,
        "LC R": rb,
        "LCS R": c["D4"] * rb,
        "LCI R": c["D3"] * rb,
        "Sigma Rbarra/d2": rb / c["d2"],
    }
    limites["Puntos fuera Xbarra"] = int(((resumen["mean"] > limites["LCS Xbarra"]) | (resumen["mean"] < limites["LCI Xbarra"])).sum())
    limites["Puntos fuera R"] = int(((resumen["R"] > limites["LCS R"]) | (resumen["R"] < limites["LCI R"])).sum())
    return resumen, limites


def calcular_i_mr(serie):
    s = convertir_a_numerica(serie).reset_index(drop=True)
    if len(s) < 3:
        return None
    mr = s.diff().abs().dropna()
    mr_barra = float(mr.mean())
    if mr_barra == 0 or pd.isna(mr_barra):
        return None
    media = float(s.mean())
    limites = {
        "LC I": media,
        "LCS I": media + 2.66 * mr_barra,
        "LCI I": media - 2.66 * mr_barra,
        "MR barra": mr_barra,
        "LC MR": mr_barra,
        "LCS MR": 3.268 * mr_barra,
        "LCI MR": 0.0,
    }
    limites["Puntos fuera I"] = int(((s > limites["LCS I"]) | (s < limites["LCI I"])).sum())
    limites["Puntos fuera MR"] = int(((mr > limites["LCS MR"]) | (mr < limites["LCI MR"])).sum())
    return s, mr, limites


def reglas_shewhart(valores, lc, lcs, lci):
    s = convertir_a_numerica(valores).reset_index(drop=True)
    if len(s) == 0:
        return pd.DataFrame(columns=["Regla", "Resultado"])
    fuera = int(((s > lcs) | (s < lci)).sum())
    arriba = (s > lc).astype(int)
    abajo = (s < lc).astype(int)
    max_racha = 0
    actual = 0
    lado_anterior = None
    for a, b in zip(arriba, abajo):
        lado = "arriba" if a else "abajo" if b else "centro"
        if lado != "centro" and lado == lado_anterior:
            actual += 1
        else:
            actual = 1
            lado_anterior = lado
        max_racha = max(max_racha, actual)
    tendencias = 0
    dif = s.diff().dropna()
    if len(dif) >= 5:
        for i in range(len(dif) - 4):
            v = dif.iloc[i:i + 5]
            if (v > 0).all() or (v < 0).all():
                tendencias += 1
    return pd.DataFrame([
        {"Regla": "Puntos fuera de límites", "Resultado": fuera},
        {"Regla": "Mayor racha a un lado de LC", "Resultado": int(max_racha)},
        {"Regla": "Tendencias de 6 puntos", "Resultado": int(tendencias)},
    ])


def calcular_p(df, col_def, col_n):
    temp = pd.DataFrame({"def": pd.to_numeric(obtener_columna(df, col_def), errors="coerce"), "n": pd.to_numeric(obtener_columna(df, col_n), errors="coerce")}).dropna()
    temp = temp[(temp["n"] > 0) & (temp["def"] >= 0) & (temp["def"] <= temp["n"])]
    if temp.empty:
        return None
    pbar = float(temp["def"].sum() / temp["n"].sum())
    temp["p"] = temp["def"] / temp["n"]
    temp["LCS"] = (pbar + 3 * np.sqrt(pbar * (1 - pbar) / temp["n"])).clip(upper=1)
    temp["LCI"] = (pbar - 3 * np.sqrt(pbar * (1 - pbar) / temp["n"])).clip(lower=0)
    limites = {"p barra": pbar, "Defectuosos totales": int(temp["def"].sum()), "Inspeccionados totales": int(temp["n"].sum()), "% no conforme": pbar * 100}
    limites["Puntos fuera"] = int(((temp["p"] > temp["LCS"]) | (temp["p"] < temp["LCI"])).sum())
    return temp, limites


def calcular_np(df, col_def, col_n):
    temp = pd.DataFrame({"def": pd.to_numeric(obtener_columna(df, col_def), errors="coerce"), "n": pd.to_numeric(obtener_columna(df, col_n), errors="coerce")}).dropna()
    temp = temp[(temp["n"] > 0) & (temp["def"] >= 0) & (temp["def"] <= temp["n"])]
    if temp.empty:
        return None
    n_usado = float(temp["n"].mean())
    pbar = float(temp["def"].sum() / temp["n"].sum())
    npbar = n_usado * pbar
    sigma = np.sqrt(n_usado * pbar * (1 - pbar))
    limites = {"n usado": n_usado, "p barra": pbar, "LC np": npbar, "LCS np": npbar + 3 * sigma, "LCI np": max(0, npbar - 3 * sigma)}
    limites["Puntos fuera"] = int(((temp["def"] > limites["LCS np"]) | (temp["def"] < limites["LCI np"])).sum())
    return temp, limites


def calcular_c(df, col_defectos):
    temp = pd.DataFrame({"c": pd.to_numeric(obtener_columna(df, col_defectos), errors="coerce")}).dropna()
    temp = temp[temp["c"] >= 0]
    if temp.empty:
        return None
    cbar = float(temp["c"].mean())
    limites = {"LC c": cbar, "LCS c": cbar + 3 * np.sqrt(cbar), "LCI c": max(0, cbar - 3 * np.sqrt(cbar))}
    limites["Puntos fuera"] = int(((temp["c"] > limites["LCS c"]) | (temp["c"] < limites["LCI c"])).sum())
    return temp, limites


def calcular_u(df, col_defectos, col_unidades):
    temp = pd.DataFrame({"def": pd.to_numeric(obtener_columna(df, col_defectos), errors="coerce"), "n": pd.to_numeric(obtener_columna(df, col_unidades), errors="coerce")}).dropna()
    temp = temp[(temp["def"] >= 0) & (temp["n"] > 0)]
    if temp.empty:
        return None
    ubar = float(temp["def"].sum() / temp["n"].sum())
    temp["u"] = temp["def"] / temp["n"]
    temp["LCS"] = ubar + 3 * np.sqrt(ubar / temp["n"])
    temp["LCI"] = (ubar - 3 * np.sqrt(ubar / temp["n"])).clip(lower=0)
    limites = {"u barra": ubar, "Defectos totales": int(temp["def"].sum()), "Unidades totales": int(temp["n"].sum())}
    limites["Puntos fuera"] = int(((temp["u"] > temp["LCS"]) | (temp["u"] < temp["LCI"])).sum())
    return temp, limites


def curva_oc(n, c, pasos=151):
    ps = np.linspace(0, 0.25, pasos)
    return pd.DataFrame({"Fracción defectuosa": ps, "Probabilidad de aceptación": [float(stats.binom.cdf(min(c, n), n, p)) for p in ps]})


# ============================================================
# INTERFAZ Y GRÁFICOS
# ============================================================

def importar_librerias_app():
    if FALTAN_APP:
        return None, None
    import streamlit as st
    import plotly.graph_objects as go
    return st, go


def aplicar_estilo_visual(st):
    st.markdown(
        """
<style>
@keyframes entradaSuave {from {opacity:0; transform:translateY(16px) scale(.985)} to {opacity:1; transform:translateY(0) scale(1)}}
@keyframes flotar {0%,100% {transform:translateY(0)} 50% {transform:translateY(-5px)}}
@keyframes brillo {0%,100% {box-shadow:0 12px 30px rgba(15,23,42,.08)} 50% {box-shadow:0 20px 48px rgba(14,165,233,.18)}}
@keyframes destello {0% {background-position:-360px 0} 100% {background-position:360px 0}}
@keyframes puntoVivo {0%,100% {transform:scale(1); opacity:.68} 50% {transform:scale(1.35); opacity:1}}
html, body, [class*="css"] {font-family:Arial, Helvetica, sans-serif;}
.stApp {
    background:
        radial-gradient(circle at 8% 4%, rgba(14,165,233,.20), transparent 30%),
        radial-gradient(circle at 94% 8%, rgba(168,85,247,.18), transparent 31%),
        radial-gradient(circle at 50% 100%, rgba(34,197,94,.12), transparent 30%),
        linear-gradient(135deg,#f8fafc 0%,#eef2ff 42%,#fdf4ff 72%,#f8fafc 100%);
}
.block-container {padding-top:1.20rem; padding-bottom:2.4rem; max-width:1380px; animation:entradaSuave .45s ease both;}
section[data-testid="stSidebar"] {background:linear-gradient(180deg,#111827 0%,#1e1b4b 54%,#0f172a 100%); border-right:1px solid rgba(255,255,255,.12);}
section[data-testid="stSidebar"] * {color:#f8fafc!important;}
div[role="radiogroup"] label {border-radius:20px!important; padding:.42rem .65rem!important; margin:.12rem 0!important; border:1px solid rgba(255,255,255,.08)!important; transition:all .18s ease!important;}
div[role="radiogroup"] label:hover {background:rgba(255,255,255,.13)!important; transform:translateX(5px) scale(1.01);}
.app-title {position:relative; overflow:hidden; padding:1.35rem 1.55rem; border-radius:34px; background:linear-gradient(135deg,rgba(255,255,255,.95),rgba(240,249,255,.90)); border:1px solid rgba(148,163,184,.25); box-shadow:0 20px 48px rgba(15,23,42,.10); margin-bottom:1.15rem; backdrop-filter:blur(14px); animation:entradaSuave .55s ease both;}
.app-title:before {content:""; position:absolute; width:190px; height:190px; right:-55px; top:-85px; background:linear-gradient(135deg,#38bdf8,#a78bfa,#fb7185); border-radius:50%; opacity:.18;}
.app-title:after {content:""; position:absolute; inset:0; background:linear-gradient(110deg,transparent,rgba(255,255,255,.58),transparent); background-size:360px 100%; animation:destello 4.6s linear infinite; pointer-events:none;}
.creator-tag {margin-top:.5rem; font-size:.78rem; color:#64748b; font-weight:600;}
.soft-card {min-height:108px; padding:1.05rem; border-radius:26px; background:linear-gradient(145deg,rgba(255,255,255,.98),rgba(248,250,252,.91)); border:1px solid rgba(148,163,184,.22); box-shadow:0 14px 32px rgba(15,23,42,.08); transition:all .2s ease; animation:entradaSuave .45s ease both, flotar 6s ease-in-out infinite;}
.soft-card:hover {transform:translateY(-6px) scale(1.015); animation:brillo 1.7s ease infinite; border-color:rgba(14,165,233,.42);}
.mini-label {margin:0; color:#64748b; font-size:.73rem; letter-spacing:.055em; text-transform:uppercase; font-weight:800;}
.big-number {margin:.34rem 0 0 0; background:linear-gradient(90deg,#0f172a,#1d4ed8,#7c3aed); -webkit-background-clip:text; color:transparent; font-size:1.42rem; font-weight:900; line-height:1.1;}
.ok-box,.warn-box,.bad-box,.info-box {padding:1rem; border-radius:20px; margin:.45rem 0 .75rem 0; border:1px solid transparent; box-shadow:0 10px 24px rgba(15,23,42,.06); animation:entradaSuave .35s ease both;}
.ok-box {background:#ecfdf5; color:#065f46; border-color:#a7f3d0}.warn-box {background:#fffbeb; color:#92400e; border-color:#fde68a}.bad-box {background:#fef2f2; color:#991b1b; border-color:#fecaca}.info-box {background:#eff6ff; color:#1e3a8a; border-color:#bfdbfe}
div[data-testid="stDataFrame"] {border-radius:20px; overflow:hidden; box-shadow:0 16px 34px rgba(15,23,42,.08); animation:entradaSuave .45s ease both;}
.live-badge {display:inline-flex; gap:.4rem; padding:.48rem .8rem; border-radius:999px; background:#ecfeff; color:#155e75; border:1px solid #a5f3fc; font-weight:900; animation:brillo 1.6s ease-in-out infinite;}
.live-badge:before {content:'●'; color:#22c55e; animation:puntoVivo 1s ease-in-out infinite;}
button {border-radius:18px!important; transition:all .18s ease!important; font-weight:700!important;}
button:hover {transform:translateY(-2px)!important; box-shadow:0 12px 28px rgba(37,99,235,.15)!important;}
</style>
""",
        unsafe_allow_html=True,
    )


def encabezado(st, titulo, subtitulo):
    st.markdown(
        f"""
<div class="app-title">
    <h1 style="margin:0; color:#0f172a; font-size:2.05rem;">{titulo}</h1>
    <p style="margin:.25rem 0 0 0; color:#475569; font-size:1rem;">{subtitulo}</p>
    <p class="creator-tag">Creado por Jerson Andrés López Wilches</p>
</div>
""",
        unsafe_allow_html=True,
    )


def caja_estado(st, tipo, texto):
    clase = {"ok": "ok-box", "alerta": "warn-box", "error": "bad-box", "info": "info-box"}.get(tipo, "info-box")
    st.markdown(f"<div class='{clase}'>{texto}</div>", unsafe_allow_html=True)


def tarjetas(st, datos: dict, titulo: str):
    st.markdown(f"### {titulo}")
    claves = list(datos.keys())
    for i in range(0, len(claves), 4):
        cols = st.columns(min(4, len(claves) - i))
        for j, k in enumerate(claves[i:i + 4]):
            with cols[j]:
                st.markdown(f"<div class='soft-card'><p class='mini-label'>{k}</p><p class='big-number'>{fmt(datos[k])}</p></div>", unsafe_allow_html=True)


def tabla(st, datos: dict, titulo="Resultados"):
    st.markdown(f"### {titulo}")
    df = pd.DataFrame([redondear_dict(datos, 3)]).T.reset_index()
    df.columns = ["Indicador", "Valor"]
    st.dataframe(df, use_container_width=True, hide_index=True)


def estilo_fig(fig, titulo, subtitulo=""):
    texto = titulo if not subtitulo else f"{titulo}<br><sup>{subtitulo}</sup>"
    fig.update_layout(
        title={"text": texto, "x": 0.03, "xanchor": "left", "font": {"size": 22}},
        template="plotly_white",
        height=510,
        margin=dict(l=48, r=34, t=90, b=52),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        hovermode="x unified",
        font=dict(size=13, family="Arial"),
        paper_bgcolor="rgba(255,255,255,0)",
        plot_bgcolor="rgba(255,255,255,.98)",
    )
    fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor="rgba(148,163,184,.20)", zeroline=False, showline=True, linecolor="rgba(15,23,42,.35)")
    fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor="rgba(148,163,184,.20)", zeroline=False, showline=True, linecolor="rgba(15,23,42,.35)")
    return fig


def hline(fig, y, nombre, dash="dash"):
    fig.add_hline(y=y, line_dash=dash, line_width=2.8, annotation_text=nombre, annotation_position="top left")


def vline(fig, x, nombre, dash="dash"):
    fig.add_vline(x=x, line_dash=dash, line_width=2.8, annotation_text=nombre, annotation_position="top")


def grafico_corrida(go, serie, titulo="Gráfico de corrida"):
    s = convertir_a_numerica(serie).reset_index(drop=True)
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=s.index + 1, y=s, mode="lines+markers", name="Medición", line=dict(width=3), marker=dict(size=8)))
    if len(s):
        hline(fig, float(s.mean()), "Media", "solid")
    fig.update_xaxes(title="Orden")
    fig.update_yaxes(title="Medición")
    return estilo_fig(fig, titulo, "Evalúa tendencia, saltos, ciclos y comportamiento temporal")


def grafico_qq(go, serie):
    s = convertir_a_numerica(serie)
    fig = go.Figure()
    if len(s) < 3:
        return estilo_fig(fig, "Q-Q normal", "Datos insuficientes")
    osm, osr = stats.probplot(s, dist="norm", fit=False)
    pendiente, intercepto, _ = stats.probplot(s, dist="norm", fit=True)[1]
    xline = np.array([min(osm), max(osm)])
    yline = intercepto + pendiente * xline
    fig.add_trace(go.Scatter(x=osm, y=osr, mode="markers", name="Datos", marker=dict(size=9)))
    fig.add_trace(go.Scatter(x=xline, y=yline, mode="lines", name="Referencia normal", line=dict(width=4)))
    fig.update_xaxes(title="Cuantiles teóricos")
    fig.update_yaxes(title="Cuantiles observados")
    return estilo_fig(fig, "Gráfico Q-Q de normalidad", "Puntos cercanos a la línea indican ajuste normal")


def grafico_box(go, serie):
    s = convertir_a_numerica(serie)
    fig = go.Figure()
    fig.add_trace(go.Box(y=s, boxpoints="all", jitter=0.28, pointpos=0, name="Datos", marker=dict(size=7)))
    fig.update_yaxes(title="Medición")
    return estilo_fig(fig, "Boxplot de mediciones", "Visualiza dispersión, mediana y valores atípicos")


def grafico_hist_capacidad(go, serie, lie, lse, vn=None):
    s = convertir_a_numerica(serie)
    media = float(s.mean())
    sigma = float(s.std(ddof=1)) if len(s) > 1 else np.nan
    fig = go.Figure()
    fig.add_trace(go.Histogram(x=s, nbinsx=26, name="Datos observados", opacity=0.78, marker=dict(line=dict(width=1, color="rgba(15,23,42,.25)"))))
    if len(s) > 1 and sigma > 0:
        x_min = min(float(s.min()), lie) - sigma
        x_max = max(float(s.max()), lse) + sigma
        x = np.linspace(x_min, x_max, 350)
        y = stats.norm.pdf(x, media, sigma)
        escala = len(s) * (x_max - x_min) / 26
        fig.add_trace(go.Scatter(x=x, y=y * escala, mode="lines", name="Aproximación normal", line=dict(width=4)))
    vline(fig, lie, "LIE")
    vline(fig, lse, "LSE")
    vline(fig, media, "Media", "solid")
    if vn is not None:
        vline(fig, vn, "VN", "dot")
    fig.update_xaxes(title="Variable de calidad")
    fig.update_yaxes(title="Frecuencia")
    return estilo_fig(fig, "Capacidad del proceso", "Histograma, normal aproximada, LIE, LSE y VN")


def grafico_escenario_mejora(go, media_actual, sigma_actual, media_nueva, sigma_objetivo, lie, lse, vn=None):
    if sigma_actual <= 0 or sigma_objetivo <= 0:
        return None
    minimo = min(lie, media_actual - 4 * sigma_actual, media_nueva - 4 * sigma_objetivo)
    maximo = max(lse, media_actual + 4 * sigma_actual, media_nueva + 4 * sigma_objetivo)
    x = np.linspace(minimo, maximo, 500)
    y_actual = stats.norm.pdf(x, media_actual, sigma_actual)
    y_nueva = stats.norm.pdf(x, media_nueva, sigma_objetivo)
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=x, y=y_actual, mode="lines", name="Proceso actual", line=dict(width=4)))
    fig.add_trace(go.Scatter(x=x, y=y_nueva, mode="lines", name="Escenario mejorado", line=dict(width=4, dash="dash")))
    vline(fig, lie, "LIE")
    vline(fig, lse, "LSE")
    vline(fig, media_actual, "Media actual", "solid")
    vline(fig, media_nueva, "Nueva media", "dot")
    if vn is not None:
        vline(fig, vn, "VN", "dashdot")
    fig.update_xaxes(title="Variable de calidad")
    fig.update_yaxes(title="Densidad normal")
    return estilo_fig(fig, "Escenario de mejora", "Compara el proceso actual contra sigma objetivo y nueva media")


def grafico_xbar_r(go, df, col_valor, col_subgrupo):
    calc = calcular_xbar_r(df, col_valor, col_subgrupo)
    if calc is None:
        return None, None, None
    resumen, lim = calc
    fig_x = go.Figure()
    fig_x.add_trace(go.Scatter(x=resumen.index.astype(str), y=resumen["mean"], mode="lines+markers", name="Media del subgrupo", line=dict(width=3), marker=dict(size=8)))
    hline(fig_x, lim["LC Xbarra"], "LC X̄", "solid")
    hline(fig_x, lim["LCS Xbarra"], "LCS X̄")
    hline(fig_x, lim["LCI Xbarra"], "LCI X̄")
    fig_x.update_xaxes(title="Subgrupo")
    fig_x.update_yaxes(title="Media")
    fig_x = estilo_fig(fig_x, "Carta X-barra", "Control de la media del proceso")
    fig_r = go.Figure()
    fig_r.add_trace(go.Scatter(x=resumen.index.astype(str), y=resumen["R"], mode="lines+markers", name="Rango", line=dict(width=3), marker=dict(size=8)))
    hline(fig_r, lim["LC R"], "LC R", "solid")
    hline(fig_r, lim["LCS R"], "LCS R")
    hline(fig_r, lim["LCI R"], "LCI R")
    fig_r.update_xaxes(title="Subgrupo")
    fig_r.update_yaxes(title="Rango")
    fig_r = estilo_fig(fig_r, "Carta R", "Control de la variabilidad dentro del subgrupo")
    return fig_x, fig_r, lim


def grafico_i_mr(go, serie):
    calc = calcular_i_mr(serie)
    if calc is None:
        return None, None, None
    s, mr, lim = calc
    fig_i = go.Figure()
    fig_i.add_trace(go.Scatter(x=s.index + 1, y=s, mode="lines+markers", name="Individual", line=dict(width=3), marker=dict(size=8)))
    hline(fig_i, lim["LC I"], "LC I", "solid")
    hline(fig_i, lim["LCS I"], "LCS I")
    hline(fig_i, lim["LCI I"], "LCI I")
    fig_i.update_xaxes(title="Observación")
    fig_i.update_yaxes(title="Valor individual")
    fig_i = estilo_fig(fig_i, "Carta I", "Valores individuales")
    fig_mr = go.Figure()
    fig_mr.add_trace(go.Scatter(x=mr.index + 1, y=mr, mode="lines+markers", name="Rango móvil", line=dict(width=3), marker=dict(size=8)))
    hline(fig_mr, lim["LC MR"], "LC MR", "solid")
    hline(fig_mr, lim["LCS MR"], "LCS MR")
    hline(fig_mr, lim["LCI MR"], "LCI MR")
    fig_mr.update_xaxes(title="Observación")
    fig_mr.update_yaxes(title="Rango móvil")
    fig_mr = estilo_fig(fig_mr, "Carta MR", "Rango móvil entre observaciones consecutivas")
    return fig_i, fig_mr, lim


def grafico_lineas_atributos(go, x, y, lc=None, lcs=None, lci=None, titulo="Carta de control", subtitulo=""):
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=x, y=y, mode="lines+markers", name="Valor", line=dict(width=3), marker=dict(size=8)))
    if lc is not None:
        hline(fig, lc, "LC", "solid")
    if lcs is not None:
        if np.isscalar(lcs):
            hline(fig, lcs, "LCS")
        else:
            fig.add_trace(go.Scatter(x=x, y=lcs, mode="lines", name="LCS", line=dict(width=3, dash="dash")))
    if lci is not None:
        if np.isscalar(lci):
            hline(fig, lci, "LCI")
        else:
            fig.add_trace(go.Scatter(x=x, y=lci, mode="lines", name="LCI", line=dict(width=3, dash="dash")))
    fig.update_xaxes(title="Muestra")
    fig.update_yaxes(title="Valor de control")
    return estilo_fig(fig, titulo, subtitulo)


def grafico_oc(go, datos, aql=None, ltpd=None):
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=datos["Fracción defectuosa"], y=datos["Probabilidad de aceptación"], mode="lines", name="Curva OC", line=dict(width=4)))
    if aql is not None:
        vline(fig, aql, "AQL / NAC")
    if ltpd is not None:
        vline(fig, ltpd, "LTPD / NRC")
    fig.update_xaxes(title="Fracción defectuosa")
    fig.update_yaxes(title="Probabilidad de aceptación")
    return estilo_fig(fig, "Curva característica de operación", "Riesgo del productor y consumidor")


# ============================================================
# PANTALLAS
# ============================================================

def pantalla_inicio(st):
    tarjetas(st, {"Estado": "Listo", "Enfoque": "Control estadístico", "Modo": "Interactivo", "Creador": "Jerson López"}, "Panel principal")
    st.markdown("### Ruta de análisis")
    c1, c2, c3 = st.columns(3)
    with c1:
        caja_estado(st, "info", "1. Carga datos y define la variable crítica de calidad.")
        caja_estado(st, "info", "2. Valida normalidad, independencia, homocedasticidad y atípicos.")
    with c2:
        caja_estado(st, "info", "3. Selecciona la carta correcta: I-MR, X-barra/R, p, np, c o u.")
        caja_estado(st, "info", "4. Evalúa Cp, Cpk, Pp, Ppk, PNC y escenarios de mejora.")
    with c3:
        caja_estado(st, "info", "5. Revisa no conformes y riesgos por LIE o LSE.")
        caja_estado(st, "info", "6. Exporta reporte y usa monitoreo para seguimiento continuo.")
    st.markdown("### Diccionario de conceptos")
    dic = pd.DataFrame({
        "Sigla": ["VN", "LIE", "LSE", "Cp", "CPU", "CPL", "Cpk", "PNC", "ARL", "ATS"],
        "Nombre completo": ["Valor nominal u objetivo", "Límite inferior de especificación", "Límite superior de especificación", "Capacidad potencial", "Capacidad hacia LSE", "Capacidad hacia LIE", "Capacidad real", "Producto no conforme", "Promedio de muestras hasta señal", "Tiempo promedio hasta señal"],
    })
    st.dataframe(dic, use_container_width=True, hide_index=True)


def pantalla_cargar_datos(st):
    st.markdown("### Cargar datos")
    archivo = st.file_uploader("Sube un archivo CSV o Excel", type=["csv", "xlsx"])
    if archivo is not None:
        try:
            df = cargar_datos_desde_archivo(archivo)
            st.session_state["df"] = df
            st.session_state["df_original"] = df.copy()
            st.success("Datos cargados correctamente.")
        except Exception as e:
            st.error(f"Error al cargar archivo: {e}")
    texto = st.text_area("O pega datos en formato CSV", height=130, placeholder="subgrupo,medicion\n1,10.1\n1,9.9")
    if st.button("Cargar datos manuales"):
        try:
            df = normalizar_nombres_columnas(pd.read_csv(StringIO(texto)))
            st.session_state["df"] = df
            st.session_state["df_original"] = df.copy()
            st.success("Datos manuales cargados.")
        except Exception as e:
            st.error(f"No se pudieron cargar los datos: {e}")

    with st.expander("Ingresar datos en una tabla sencilla"):
        caja_estado(
            st,
            "info",
            "Escribe o pega los datos como en Excel. Puedes agregar filas, cambiar nombres de columnas y borrar columnas que no uses."
        )
        guardar_estado_si_no_existe(st, "datos_tabla_simple", pd.DataFrame({
            "subgrupo": [1, 1, 2, 2],
            "medicion": [10.1, 9.9, 10.3, 10.0],
        }))
        datos_tabla = st.data_editor(
            st.session_state["datos_tabla_simple"],
            num_rows="dynamic",
            use_container_width=True,
            key="editor_datos_tabla_simple",
        )
        c1, c2 = st.columns([1, 3])
        with c1:
            if st.button("Usar tabla sencilla"):
                try:
                    df = normalizar_nombres_columnas(datos_tabla.dropna(how="all"))
                    if df.empty:
                        st.error("La tabla está vacía.")
                    else:
                        st.session_state["datos_tabla_simple"] = df.copy()
                        st.session_state["df"] = df
                        st.session_state["df_original"] = df.copy()
                        st.success("Datos de la tabla cargados.")
                except Exception as e:
                    st.error(f"No se pudieron cargar los datos de la tabla: {e}")
        with c2:
            st.caption("Esta opción es útil cuando no tienes archivo. Las columnas recomendadas son subgrupo y medicion para cartas X-barra/R.")

    df = st.session_state.get("df")
    if dataframe_valido(df):
        tarjetas(st, {"Filas": len(df), "Columnas": len(df.columns), "Numéricas": len(columnas_numericas(df)), "Faltantes": int(df.isna().sum().sum())}, "Resumen del archivo")
        st.dataframe(df.head(40), use_container_width=True, hide_index=True)
        with st.expander("Convertir formato ancho a formato largo"):
            col_sub = st.selectbox("Columna que identifica el subgrupo", df.columns.tolist())
            if st.button("Convertir"):
                st.session_state["df"] = convertir_ancho_a_largo(df, col_sub)
                st.success("Datos convertidos a formato largo.")
                st.dataframe(st.session_state["df"], use_container_width=True, hide_index=True)
        st.markdown("### Resumen estadístico")
        st.dataframe(df.describe(include="all").T.reset_index().rename(columns={"index": "Variable"}), use_container_width=True, hide_index=True)


def pantalla_supuestos(st, go):
    df = st.session_state.get("df")
    if not dataframe_valido(df):
        st.warning("Primero carga datos.")
        return
    nums = columnas_numericas(df)
    if not nums:
        st.error("No hay columnas numéricas.")
        return

    col = selectbox_persistente(st, "Variable de medición", nums, "sup_variable", nums[0])
    s = convertir_a_numerica(obtener_columna(df, col))
    tarjetas(st, resumen_descriptivo(s), "Resumen de la variable")
    tab1, tab2, tab3, tab4 = st.tabs(["Normalidad", "Independencia", "Homocedasticidad", "Atípicos"])

    with tab1:
        norm = evaluar_normalidad(s)
        estado, texto = diagnostico_normalidad(norm)
        caja_estado(st, "ok" if estado == "Cumple" else "alerta" if estado == "No cumple" else "info", texto)
        st.dataframe(norm, use_container_width=True, hide_index=True)
        a, b = st.columns(2)
        with a:
            st.plotly_chart(grafico_qq(go, s), use_container_width=True, key="supuestos_qq")
        with b:
            st.plotly_chart(grafico_box(go, s), use_container_width=True, key="box_supuestos_normalidad")

    with tab2:
        ind = evaluar_independencia(s)
        tarjetas(st, ind, "Independencia temporal")
        st.plotly_chart(grafico_corrida(go, s), use_container_width=True, key="supuestos_corrida")

    with tab3:
        opciones_grupo = ["Ninguna"] + df.columns.tolist()
        grupo = selectbox_persistente(st, "Columna de grupo", opciones_grupo, "sup_grupo", "Ninguna")
        if grupo != "Ninguna":
            st.dataframe(evaluar_homocedasticidad(df, col, grupo), use_container_width=True, hide_index=True)
        else:
            caja_estado(st, "info", "Selecciona una columna de grupo para evaluar homocedasticidad.")

    with tab4:
        li, ls, atip = detectar_atipicos_iqr(s)
        tarjetas(st, {"Límite IQR inferior": li, "Límite IQR superior": ls, "Atípicos": len(atip)}, "Detección IQR")
        st.plotly_chart(grafico_box(go, s), use_container_width=True, key="box_supuestos_atipicos")
        if len(atip) > 0:
            st.dataframe(atip.reset_index(), use_container_width=True, hide_index=True)


def pantalla_control(st, go):
    df = st.session_state.get("df")
    if not dataframe_valido(df):
        st.warning("Primero carga datos.")
        return
    nums = columnas_numericas(df)
    if not nums:
        st.error("No hay columnas numéricas.")
        return

    tipo = radio_persistente(
        st,
        "Carta de control",
        ["I-MR", "X-barra y R", "p", "np", "c", "u"],
        "ctrl_tipo",
        "I-MR",
        horizontal=True,
    )

    if tipo == "I-MR":
        col = selectbox_persistente(st, "Variable continua", nums, "ctrl_imr_variable", nums[0])
        fig_i, fig_mr, lim = grafico_i_mr(go, obtener_columna(df, col))
        if fig_i is None:
            st.error("No se pudo calcular I-MR. Revisa que haya variación y al menos 3 datos.")
            return
        tarjetas(st, lim, "Límites I-MR")
        t1, t2, t3 = st.tabs(["Carta I", "Carta MR", "Reglas"])
        with t1:
            st.plotly_chart(fig_i, use_container_width=True, key="control_imr_i")
        with t2:
            st.plotly_chart(fig_mr, use_container_width=True, key="control_imr_mr")
        with t3:
            st.dataframe(reglas_shewhart(convertir_a_numerica(obtener_columna(df, col)), lim["LC I"], lim["LCS I"], lim["LCI I"]), use_container_width=True, hide_index=True)

    elif tipo == "X-barra y R":
        col = selectbox_persistente(st, "Variable continua", nums, "ctrl_xbar_variable", nums[0])
        sub = selectbox_persistente(st, "Subgrupo", df.columns.tolist(), "ctrl_xbar_subgrupo", df.columns.tolist()[0])
        fig_x, fig_r, lim = grafico_xbar_r(go, df, col, sub)
        if fig_x is None:
            st.error("No se pudo calcular X-barra/R. Revisa que cada subgrupo tenga de 2 a 10 datos.")
            return
        tarjetas(st, lim, "Límites X-barra/R")
        calc = calcular_xbar_r(df, col, sub)
        resumen = calc[0]
        t1, t2, t3, t4 = st.tabs(["Carta X-barra", "Carta R", "Subgrupos", "Reglas"])
        with t1:
            st.plotly_chart(fig_x, use_container_width=True, key="control_xbar")
        with t2:
            st.plotly_chart(fig_r, use_container_width=True, key="control_r")
        with t3:
            st.dataframe(resumen.reset_index(), use_container_width=True, hide_index=True)
        with t4:
            st.dataframe(reglas_shewhart(resumen["mean"], lim["LC Xbarra"], lim["LCS Xbarra"], lim["LCI Xbarra"]), use_container_width=True, hide_index=True)

    elif tipo == "p":
        dcol = selectbox_persistente(st, "Defectuosos", nums, "ctrl_p_defectuosos", nums[0])
        ncol = selectbox_persistente(st, "Inspeccionados", nums, "ctrl_p_inspeccionados", nums[0])
        calc = calcular_p(df, dcol, ncol)
        if calc is None:
            st.error("No se pudo calcular p.")
            return
        temp, lim = calc
        tarjetas(st, lim, "Carta p")
        st.plotly_chart(grafico_lineas_atributos(go, temp.index + 1, temp["p"], lim["p barra"], temp["LCS"], temp["LCI"], "Carta p", "Proporción no conforme"), use_container_width=True, key="control_p")

    elif tipo == "np":
        dcol = selectbox_persistente(st, "Defectuosos", nums, "ctrl_np_defectuosos", nums[0])
        ncol = selectbox_persistente(st, "Inspeccionados", nums, "ctrl_np_inspeccionados", nums[0])
        calc = calcular_np(df, dcol, ncol)
        if calc is None:
            st.error("No se pudo calcular np.")
            return
        temp, lim = calc
        tarjetas(st, lim, "Carta np")
        st.plotly_chart(grafico_lineas_atributos(go, temp.index + 1, temp["def"], lim["LC np"], lim["LCS np"], lim["LCI np"], "Carta np", "Número de unidades no conformes"), use_container_width=True, key="control_np")

    elif tipo == "c":
        ccol = selectbox_persistente(st, "Defectos", nums, "ctrl_c_defectos", nums[0])
        calc = calcular_c(df, ccol)
        if calc is None:
            st.error("No se pudo calcular c.")
            return
        temp, lim = calc
        tarjetas(st, lim, "Carta c")
        st.plotly_chart(grafico_lineas_atributos(go, temp.index + 1, temp["c"], lim["LC c"], lim["LCS c"], lim["LCI c"], "Carta c", "Número de defectos"), use_container_width=True, key="control_c")

    else:
        dcol = selectbox_persistente(st, "Defectos", nums, "ctrl_u_defectos", nums[0])
        ncol = selectbox_persistente(st, "Unidades inspeccionadas", nums, "ctrl_u_unidades", nums[0])
        calc = calcular_u(df, dcol, ncol)
        if calc is None:
            st.error("No se pudo calcular u.")
            return
        temp, lim = calc
        tarjetas(st, lim, "Carta u")
        st.plotly_chart(grafico_lineas_atributos(go, temp.index + 1, temp["u"], lim["u barra"], temp["LCS"], temp["LCI"], "Carta u", "Defectos por unidad"), use_container_width=True, key="control_u")


def pantalla_capacidad(st, go):
    df = st.session_state.get("df")
    if not dataframe_valido(df):
        st.warning("Primero carga datos.")
        return

    nums = columnas_numericas(df)
    if not nums:
        st.error("No hay columnas numéricas.")
        return

    st.markdown("### Análisis de capacidad del proceso")
    caja_estado(
        st,
        "info",
        "Los valores quedan guardados aunque cambies de módulo. Este apartado no incluye boxplot para hacerlo más liviano."
    )

    variable = selectbox_persistente(st, "Variable crítica de calidad", nums, "cap_variable", nums[0])
    serie = convertir_a_numerica(obtener_columna(df, variable))
    if len(serie) < 2:
        st.error("La variable seleccionada necesita al menos 2 datos numéricos.")
        return

    media_base = float(serie.mean())
    sigma_base = float(serie.std(ddof=1)) if len(serie) > 1 else 1.0
    if pd.isna(sigma_base) or sigma_base <= 0:
        sigma_base = 1.0

    guardar_estado_si_no_existe(st, "cap_vn", media_base)
    guardar_estado_si_no_existe(st, "cap_usar_tolerancia", True)
    guardar_estado_si_no_existe(st, "cap_tolerancia", float(max(1.0, sigma_base)))
    guardar_estado_si_no_existe(st, "cap_lie", float(serie.min()))
    guardar_estado_si_no_existe(st, "cap_lse", float(serie.max()))
    guardar_estado_si_no_existe(st, "cap_metodo_sigma", "Sigma muestral")
    guardar_estado_si_no_existe(st, "cap_sigma_historica", sigma_base)
    guardar_estado_si_no_existe(st, "cap_rbarra", 10.100)
    guardar_estado_si_no_existe(st, "cap_d2", 2.326)
    guardar_estado_si_no_existe(st, "cap_usar_media_conocida", False)
    guardar_estado_si_no_existe(st, "cap_media_conocida", media_base)
    guardar_estado_si_no_existe(st, "cap_cp_objetivo", 1.33)
    guardar_estado_si_no_existe(st, "cap_pnc_lse_permitido", 0.05)
    guardar_estado_si_no_existe(st, "cap_media_detectar", media_base)
    guardar_estado_si_no_existe(st, "cap_nsub", 5)
    guardar_estado_si_no_existe(st, "cap_intervalo", 15.0)
    guardar_estado_si_no_existe(st, "cap_potencia_objetivo", 0.90)

    col_config, col_resumen = st.columns([1.05, 0.95])

    with col_config:
        vn = numero_persistente(st, "VN | Valor nominal u objetivo", "cap_vn", media_base)
        usar_tolerancia = checkbox_persistente(
            st,
            "Calcular LIE y LSE desde VN ± tolerancia",
            "cap_usar_tolerancia",
            True,
        )

        if usar_tolerancia:
            tolerancia = numero_persistente(
                st,
                "Tolerancia",
                "cap_tolerancia",
                float(max(1.0, sigma_base)),
                min_value=0.000001,
            )
            lie = vn - tolerancia
            lse = vn + tolerancia
            st.write(f"LIE = {fmt(lie)} | LSE = {fmt(lse)}")
        else:
            lie = numero_persistente(
                st,
                "LIE | Límite inferior de especificación",
                "cap_lie",
                float(serie.min()),
            )
            lse = numero_persistente(
                st,
                "LSE | Límite superior de especificación",
                "cap_lse",
                float(serie.max()),
            )

        metodo_sigma = radio_persistente(
            st,
            "Método de sigma",
            ["Sigma muestral", "Sigma histórica", "R̄/d2"],
            "cap_metodo_sigma",
            "Sigma muestral",
            horizontal=True,
        )

        sigma_historica = None
        if metodo_sigma == "Sigma histórica":
            sigma_historica = numero_persistente(
                st,
                "Sigma histórica",
                "cap_sigma_historica",
                sigma_base,
                min_value=0.000001,
                format="%.8f",
            )
        elif metodo_sigma == "R̄/d2":
            rbarra = numero_persistente(
                st,
                "R̄ | Rango promedio",
                "cap_rbarra",
                10.100,
                min_value=0.000001,
            )
            d2 = numero_persistente(
                st,
                "d2",
                "cap_d2",
                2.326,
                min_value=0.000001,
            )
            sigma_historica = rbarra / d2
            st.write(f"Sigma estimada = {fmt(sigma_historica)}")

        usar_media = checkbox_persistente(
            st,
            "Usar media conocida",
            "cap_usar_media_conocida",
            False,
        )
        media_conocida = None
        if usar_media:
            media_conocida = numero_persistente(
                st,
                "Media conocida",
                "cap_media_conocida",
                media_base,
                format="%.8f",
            )

    with col_resumen:
        tarjetas(st, resumen_descriptivo(serie), "Resumen de la variable")
        caja_estado(st, "info", "Para revisar boxplot y atípicos entra a Supuestos del proceso → Atípicos.")

    if lie >= lse:
        st.error("LIE debe ser menor que LSE.")
        return

    resultado = calcular_capacidad(serie, lie, lse, vn, sigma_historica, media_conocida)
    if resultado is None:
        st.error("No se pudo calcular capacidad. Revisa la variabilidad, LIE y LSE.")
        return

    tarjetas(st, {
        "Cp": resultado["Cp"],
        "CPL": resultado["CPL"],
        "CPU": resultado["CPU"],
        "Cpk": resultado["Cpk"],
        "PNC estimado total %": resultado["% PNC estimado total"],
        "Estado": resultado["Estado"],
        "Centrado": resultado["Centrado"],
        "Riesgo": resultado["Riesgo principal"],
    }, "Indicadores de capacidad")

    t1, t2, t3, t4, t5 = st.tabs([
        "Gráfica",
        "Tabla completa",
        "Producto no conforme",
        "Escenario de mejora",
        "Potencia X-barra",
    ])

    with t1:
        st.plotly_chart(
            grafico_hist_capacidad(go, serie, lie, lse, vn),
            use_container_width=True,
            key="capacidad_hist",
        )

    with t2:
        tabla(st, resultado, "Capacidad completa")

    with t3:
        tarjetas(st, {
            "Fuera por LIE observado": resultado["Fuera por LIE observado"],
            "Fuera por LSE observado": resultado["Fuera por LSE observado"],
            "% PNC LIE estimado": resultado["% PNC estimado LIE"],
            "% PNC LSE estimado": resultado["% PNC estimado LSE"],
            "% PNC total estimado": resultado["% PNC estimado total"],
            "PPM estimado": resultado["PPM estimado"],
        }, "Producto no conforme")

    with t4:
        st.markdown("#### Simulación de mejora del proceso")
        cp_objetivo = numero_persistente(
            st,
            "Cp objetivo",
            "cap_cp_objetivo",
            1.33,
            min_value=0.1,
            step=0.01,
        )
        pnc_lse_permitido = numero_persistente(
            st,
            "PNC máximo permitido hacia LSE",
            "cap_pnc_lse_permitido",
            0.05,
            min_value=0.000001,
            max_value=0.5,
            step=0.001,
            format="%.8f",
        )

        sigma_objetivo = calcular_sigma_objetivo(lie, lse, cp_objetivo)
        nueva_media = media_maxima_para_pnc_lse(lse, sigma_objetivo, pnc_lse_permitido)
        capacidad_nueva = calcular_capacidad(
            pd.Series([nueva_media - sigma_objetivo, nueva_media, nueva_media + sigma_objetivo]),
            lie,
            lse,
            vn,
            sigma_objetivo,
            nueva_media,
        )

        tarjetas(st, {
            "Sigma actual": resultado["Sigma usada"],
            "Sigma objetivo": sigma_objetivo,
            "Media actual": resultado["Media usada"],
            "Nueva media sugerida": nueva_media,
            "PNC permitido hacia LSE %": pnc_lse_permitido * 100,
            "Cpk nuevo estimado": capacidad_nueva["Cpk"] if capacidad_nueva else np.nan,
            "PNC total nuevo %": capacidad_nueva["% PNC estimado total"] if capacidad_nueva else np.nan,
        }, "Resultados del escenario")

        fig_mejora = grafico_escenario_mejora(
            go,
            resultado["Media usada"],
            resultado["Sigma usada"],
            nueva_media,
            sigma_objetivo,
            lie,
            lse,
            vn,
        )
        if fig_mejora is not None:
            st.plotly_chart(
                fig_mejora,
                use_container_width=True,
                key="capacidad_escenario_mejora",
            )
        caja_estado(
            st,
            "info",
            "La curva continua representa el proceso actual. La curva punteada representa el proceso con sigma objetivo y nueva media sugerida."
        )

    with t5:
        st.markdown("#### Potencia para carta X-barra")
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            cambio = numero_persistente(
                st,
                "Media que se desea detectar",
                "cap_media_detectar",
                float(vn),
            )
        with c2:
            nsub = numero_persistente(
                st,
                "Tamaño de subgrupo",
                "cap_nsub",
                5,
                min_value=1,
                step=1,
            )
        with c3:
            intervalo = numero_persistente(
                st,
                "Intervalo entre muestras en minutos",
                "cap_intervalo",
                15.0,
                min_value=1.0,
            )
        with c4:
            zcrit = numero_persistente(
                st,
                "Z crítico",
                "cap_zcrit",
                3.0,
                min_value=0.01,
                step=0.1,
            )

        pot = potencia_xbarra(resultado["Media usada"], cambio, resultado["Sigma usada"], int(nsub), float(zcrit))
        if pot:
            pot["ATS1 minutos"] = pot["ARL1"] * intervalo if np.isfinite(pot["ARL1"]) else np.inf
            tabla(st, pot, "Potencia X-barra detallada")
            pasos = [
                f"1. Media actual = {fmt(pot['Media actual'])}.",
                f"2. Media a detectar = {fmt(pot['Media a detectar'])}.",
                f"3. Sigma usada = {fmt(pot['Sigma usada'])}.",
                f"4. Diferencia absoluta = |media a detectar - media actual| = {fmt(pot['Diferencia absoluta'])}.",
                f"5. d = diferencia absoluta / sigma usada = {fmt(pot['d'])}.",
                f"6. d√n = d × raíz del tamaño de subgrupo = {fmt(pot['d√n'])}.",
                f"7. Z crítico - d√n = {fmt(pot['Z crítico - d√n'])}.",
                f"8. β = P(no detectar el cambio) = {fmt(pot['β'])}, equivalente a {fmt(pot['% β'])}%.",
                f"9. Potencia = 1 - β = {fmt(pot['Potencia = 1 - β'])}, equivalente a {fmt(pot['% Potencia'])}%.",
                f"10. ARL1 = 1 / potencia = {fmt(pot['ARL1'])}. ATS1 = ARL1 × intervalo = {fmt(pot['ATS1 minutos'])} minutos.",
            ]
            st.markdown("#### Explicación paso a paso")
            for paso in pasos:
                caja_estado(st, "info", paso)

        objetivo_pot = numero_persistente(
            st,
            "Potencia objetivo",
            "cap_potencia_objetivo",
            0.90,
            min_value=0.01,
            max_value=0.99,
        )
        n_req = n_para_potencia(resultado["Media usada"], cambio, resultado["Sigma usada"], objetivo_pot, float(zcrit))
        if n_req:
            caja_estado(st, "ok", f"Tamaño de subgrupo sugerido: n = {n_req}")

    st.markdown("### Diagnóstico")
    for mensaje in diagnostico_capacidad(resultado):
        tipo = "ok" if "cumple" in mensaje.lower() or "no se observaron" in mensaje.lower() else "alerta"
        caja_estado(st, tipo, mensaje)


def pantalla_no_conformes(st, go):
    df = st.session_state.get("df")
    if not dataframe_valido(df):
        st.warning("Primero carga datos.")
        return
    nums = columnas_numericas(df)
    if not nums:
        st.error("No hay columnas numéricas.")
        return

    modo = radio_persistente(
        st,
        "Tipo",
        ["Por especificación", "Pareto por categoría"],
        "nc_modo",
        "Por especificación",
        horizontal=True,
    )

    if modo == "Por especificación":
        col = selectbox_persistente(st, "Variable de medición", nums, "nc_variable", nums[0])
        s = convertir_a_numerica(obtener_columna(df, col))
        guardar_estado_si_no_existe(st, "nc_vn", float(s.mean()))
        guardar_estado_si_no_existe(st, "nc_lie", float(s.min()))
        guardar_estado_si_no_existe(st, "nc_lse", float(s.max()))

        c1, c2, c3 = st.columns(3)
        with c1:
            vn = numero_persistente(st, "VN", "nc_vn", float(s.mean()))
        with c2:
            lie = numero_persistente(st, "LIE", "nc_lie", float(s.min()))
        with c3:
            lse = numero_persistente(st, "LSE", "nc_lse", float(s.max()))

        valores = pd.to_numeric(obtener_columna(df, col), errors="coerce")
        n_validos = len(valores.dropna())
        bajo = df[valores < lie]
        alto = df[valores > lse]
        total = df[(valores < lie) | (valores > lse)]
        pct_lie = len(bajo) / n_validos * 100 if n_validos else 0
        pct_lse = len(alto) / n_validos * 100 if n_validos else 0
        pct_total = len(total) / n_validos * 100 if n_validos else 0
        tarjetas(st, {
            "Evaluados": n_validos,
            "Conformes": n_validos - len(total),
            "No conformes": len(total),
            "Por LIE": len(bajo),
            "% por LIE": pct_lie,
            "Por LSE": len(alto),
            "% por LSE": pct_lse,
            "% no conforme total": pct_total,
        }, "Resumen no conforme")
        detalle_nc = pd.DataFrame([
            {"Límite": "LIE", "Condición": f"{col} < {fmt(lie)}", "No conformes": len(bajo), "% sobre evaluados": pct_lie},
            {"Límite": "LSE", "Condición": f"{col} > {fmt(lse)}", "No conformes": len(alto), "% sobre evaluados": pct_lse},
            {"Límite": "Total", "Condición": "Fuera de LIE o LSE", "No conformes": len(total), "% sobre evaluados": pct_total},
        ])
        st.markdown("### Porcentaje de no conformes por límite")
        st.dataframe(detalle_nc, use_container_width=True, hide_index=True)
        st.plotly_chart(grafico_hist_capacidad(go, s, lie, lse, vn), use_container_width=True, key="no_conformes_hist")
        t1, t2, t3 = st.tabs(["Todos", "Bajo LIE", "Sobre LSE"])
        with t1:
            st.dataframe(total, use_container_width=True, hide_index=True)
        with t2:
            st.dataframe(bajo, use_container_width=True, hide_index=True)
        with t3:
            st.dataframe(alto, use_container_width=True, hide_index=True)
    else:
        cat = selectbox_persistente(st, "Columna de categoría", df.columns.tolist(), "nc_categoria", df.columns.tolist()[0])
        conteo = df[cat].astype(str).value_counts().reset_index()
        conteo.columns = ["Categoría", "Frecuencia"]
        conteo["%"] = conteo["Frecuencia"] / conteo["Frecuencia"].sum() * 100
        conteo["% acumulado"] = conteo["%"].cumsum()
        fig = go.Figure()
        fig.add_trace(go.Bar(x=conteo["Categoría"], y=conteo["Frecuencia"], name="Frecuencia"))
        fig.add_trace(go.Scatter(x=conteo["Categoría"], y=conteo["% acumulado"], mode="lines+markers", name="% acumulado", yaxis="y2", line=dict(width=4)))
        fig.update_layout(yaxis2=dict(overlaying="y", side="right", range=[0, 100], title="% acumulado"))
        st.plotly_chart(estilo_fig(fig, "Pareto de no conformidades", "Frecuencia y porcentaje acumulado"), use_container_width=True, key="pareto_no_conformes")
        st.dataframe(conteo, use_container_width=True, hide_index=True)



def pantalla_muestreo(st, go):
    st.markdown("### Muestreo de aceptación por atributos")
    caja_estado(
        st,
        "info",
        "Diseña y valida planes simples de muestreo por atributos. Usa AQL, RQL/LTPD, riesgo del productor y riesgo del consumidor para encontrar n y c."
    )

    guardar_estado_si_no_existe(st, "muestreo_producto", "unidades")
    guardar_estado_si_no_existe(st, "muestreo_lote", 2500)
    guardar_estado_si_no_existe(st, "muestreo_unidad", "Proporción defectuosa")
    guardar_estado_si_no_existe(st, "muestreo_aql", 0.015)
    guardar_estado_si_no_existe(st, "muestreo_rql", 0.10)
    guardar_estado_si_no_existe(st, "muestreo_alfa", 0.05)
    guardar_estado_si_no_existe(st, "muestreo_beta", 0.10)
    guardar_estado_si_no_existe(st, "muestreo_distribucion", "Binomial")
    guardar_estado_si_no_existe(st, "muestreo_modo", "Crear plan")
    guardar_estado_si_no_existe(st, "muestreo_n", 52)
    guardar_estado_si_no_existe(st, "muestreo_c", 2)

    modo = radio_persistente(
        st,
        "Qué deseas hacer",
        ["Crear plan", "Validar plan existente"],
        "muestreo_modo",
        "Crear plan",
        horizontal=True,
    )

    st.markdown("#### Datos del caso")
    c1, c2, c3 = st.columns(3)
    with c1:
        producto = st.text_input(
            "Producto o unidad",
            value=st.session_state.get("muestreo_producto", "unidades"),
            key="muestreo_producto",
        )
    with c2:
        lote = numero_persistente(st, "Tamaño del lote", "muestreo_lote", 2500, min_value=1, step=1)
    with c3:
        distribucion = selectbox_persistente(
            st,
            "Distribución",
            ["Binomial", "Hipergeométrica"],
            "muestreo_distribucion",
            "Binomial",
        )

    c4, c5, c6 = st.columns(3)
    with c4:
        unidad = selectbox_persistente(
            st,
            "Unidad de calidad",
            ["Proporción defectuosa", "Porcentaje defectuoso (%)", "Defectivos por millón (DPM)"],
            "muestreo_unidad",
            "Proporción defectuosa",
        )
    with c5:
        aql_val = numero_persistente(
            st,
            "AQL | calidad aceptable",
            "muestreo_aql",
            0.015,
            min_value=0.000001,
            step=0.001,
            format="%.6f",
        )
    with c6:
        rql_val = numero_persistente(
            st,
            "RQL / LTPD | calidad rechazable",
            "muestreo_rql",
            0.10,
            min_value=0.000001,
            step=0.001,
            format="%.6f",
        )

    aql = _proporcion_desde_unidad(aql_val, unidad)
    rql = _proporcion_desde_unidad(rql_val, unidad)
    if aql <= 0 or rql <= 0 or aql >= rql:
        st.error("AQL debe ser menor que RQL/LTPD. Revisa las unidades elegidas.")
        return

    with st.expander("Opciones de riesgo", expanded=False):
        st.write("Valores comunes: alfa = 0.05 y beta = 0.10. Puedes cambiarlos si el ejercicio los da.")
        r1, r2 = st.columns(2)
        with r1:
            alfa = numero_persistente(
                st,
                "Alfa | riesgo del productor",
                "muestreo_alfa",
                0.05,
                min_value=0.000001,
                max_value=0.499999,
                step=0.005,
                format="%.6f",
            )
        with r2:
            beta = numero_persistente(
                st,
                "Beta | riesgo del consumidor",
                "muestreo_beta",
                0.10,
                min_value=0.000001,
                max_value=0.499999,
                step=0.005,
                format="%.6f",
            )

    lote_int = int(lote) if lote and lote > 0 else None

    if modo == "Crear plan":
        plan = disenar_plan_atributos_r_minitab(
            aql=aql,
            rql=rql,
            alfa=float(alfa),
            beta=float(beta),
            lote=lote_int,
            distribucion=distribucion,
        )
    else:
        v1, v2 = st.columns(2)
        with v1:
            n_val = numero_persistente(st, "n | tamaño de muestra", "muestreo_n", 52, min_value=1, step=1)
        with v2:
            c_val = numero_persistente(st, "c | número de aceptación", "muestreo_c", 2, min_value=0, step=1)
        if int(c_val) > int(n_val):
            st.error("c no puede ser mayor que n.")
            return
        plan = validar_plan_atributos(
            n=int(n_val),
            c=int(c_val),
            aql=aql,
            rql=rql,
            alfa=float(alfa),
            beta=float(beta),
            lote=lote_int,
            distribucion=distribucion,
        )

    tarjetas(st, {
        "Muestra n": plan["Muestra n"],
        "Acepta si X ≤": plan["Aceptación c"],
        "Rechaza si X ≥": plan["Rechazo r"],
        "Pa en AQL %": plan["Pa en AQL"] * 100,
        "Pa en RQL %": plan["Pa en RQL"] * 100,
        "Cumple": plan["Cumple"],
        "Distribución": plan["Distribución"],
        "Lote": plan["Lote"],
    }, "Plan generado")

    caja_estado(st, "ok" if plan["Cumple"] == "Sí" else "alerta", texto_plan_muestreo_minitab(plan, producto))

    tab1, tab2, tab3, tab4 = st.tabs(["Resultados", "Curva OC", "Comparar planes", "Texto para informe"])

    with tab1:
        resumen = pd.DataFrame([
            {"Indicador": "AQL", "Valor": _valor_a_unidad(aql, unidad), "Significado": "Nivel de calidad aceptable. Un lote con esta calidad debe aceptarse con alta probabilidad."},
            {"Indicador": "RQL / LTPD", "Valor": _valor_a_unidad(rql, unidad), "Significado": "Nivel de calidad rechazable. Un lote con esta calidad debe aceptarse con baja probabilidad."},
            {"Indicador": "Alfa objetivo", "Valor": float(alfa), "Significado": "Riesgo de rechazar un lote bueno."},
            {"Indicador": "Beta objetivo", "Valor": float(beta), "Significado": "Riesgo de aceptar un lote malo."},
            {"Indicador": "n", "Valor": plan["Muestra n"], "Significado": "Cantidad de unidades que se inspeccionan al azar."},
            {"Indicador": "c", "Valor": plan["Aceptación c"], "Significado": "Máximo de unidades no conformes permitido para aceptar el lote."},
            {"Indicador": "Pa(AQL)", "Valor": plan["Pa en AQL"], "Significado": "Probabilidad real de aceptar un lote con calidad AQL."},
            {"Indicador": "Pa(RQL)", "Valor": plan["Pa en RQL"], "Significado": "Probabilidad real de aceptar un lote con calidad RQL/LTPD."},
            {"Indicador": "Alfa real", "Valor": plan["Riesgo productor real"], "Significado": "1 - Pa(AQL)."},
            {"Indicador": "Beta real", "Valor": plan["Riesgo consumidor real"], "Significado": "Pa(RQL)."},
        ])
        st.dataframe(resumen, use_container_width=True, hide_index=True)
        caja_estado(st, "info", plan["Origen"])

    with tab2:
        datos = tabla_oc_atributos(plan)
        fig = grafico_oc(go, datos, aql, rql)
        fig.update_layout(title={"text": "Curva OC del plan de muestreo<br><sup>Probabilidad de aceptar el lote según la fracción defectuosa real</sup>"})
        st.plotly_chart(fig, use_container_width=True, key="muestreo_oc_r_minitab")
        puntos = pd.DataFrame([
            {"Punto": "AQL", "Calidad": _valor_a_unidad(aql, unidad), "Pa": plan["Pa en AQL"], "Decisión esperada": "Alta aceptación"},
            {"Punto": "RQL / LTPD", "Calidad": _valor_a_unidad(rql, unidad), "Pa": plan["Pa en RQL"], "Decisión esperada": "Baja aceptación"},
        ])
        st.dataframe(puntos, use_container_width=True, hide_index=True)

    with tab3:
        st.write("Compara cómo cambia la curva OC cuando cambias n o c.")
        comp_texto = st.text_area(
            "Planes a comparar. Escribe uno por línea como n,c",
            value=f"{plan['Muestra n']},{plan['Aceptación c']}\n{max(1, int(plan['Muestra n'] * 0.8))},{max(0, plan['Aceptación c'] - 1)}\n{int(plan['Muestra n'] * 1.2)},{plan['Aceptación c'] + 1}",
            height=95,
        )
        filas = []
        fig_comp = go.Figure()
        for linea in comp_texto.splitlines():
            partes = [x.strip() for x in linea.replace(";", ",").split(",")]
            if len(partes) != 2:
                continue
            try:
                n_i = int(float(partes[0]))
                c_i = int(float(partes[1]))
                if n_i <= 0 or c_i < 0 or c_i > n_i:
                    continue
                p_i = validar_plan_atributos(n_i, c_i, aql, rql, float(alfa), float(beta), lote_int, distribucion)
                filas.append({
                    "Plan": f"n={n_i}, c={c_i}",
                    "Pa(AQL) %": p_i["Pa en AQL"] * 100,
                    "Pa(RQL) %": p_i["Pa en RQL"] * 100,
                    "Alfa real %": p_i["Riesgo productor real"] * 100,
                    "Beta real %": p_i["Riesgo consumidor real"] * 100,
                    "Cumple": p_i["Cumple"],
                })
                oc_i = tabla_oc_atributos(p_i)
                fig_comp.add_trace(go.Scatter(x=oc_i["Fracción defectuosa"], y=oc_i["Probabilidad de aceptación"], mode="lines", name=f"n={n_i}, c={c_i}", line=dict(width=3)))
            except Exception:
                continue
        if filas:
            st.dataframe(pd.DataFrame(filas), use_container_width=True, hide_index=True)
            vline(fig_comp, aql, "AQL")
            vline(fig_comp, rql, "RQL")
            fig_comp.update_xaxes(title="Fracción defectuosa")
            fig_comp.update_yaxes(title="Probabilidad de aceptación")
            st.plotly_chart(estilo_fig(fig_comp, "Comparación de planes", "Permite ver cuál plan protege mejor al productor y al consumidor"), use_container_width=True, key="muestreo_comparacion")

    with tab4:
        st.markdown("#### Redacción automática")
        caja_estado(st, "ok", texto_plan_muestreo_minitab(plan, producto))
        st.markdown("#### Lectura sencilla")
        st.write(
            "n es la muestra que se inspecciona. c es el número máximo de unidades no conformes permitido. "
            "AQL protege al productor porque representa un lote bueno. RQL o LTPD protege al cliente porque representa un lote malo. "
            "La curva OC muestra la probabilidad de aceptar lotes con diferentes niveles reales de defectuosos."
        )


def pantalla_asistente(st):
    st.markdown("### Asistente de proyecto")
    tipo = selectbox_persistente(
        st,
        "Tipo de dato",
        ["Variable continua", "Defectuoso / no defectuoso", "Número de defectos"],
        "asis_tipo",
        "Variable continua",
    )
    sub = checkbox_persistente(st, "Tengo subgrupos racionales", "asis_subgrupos", False)
    n = numero_persistente(st, "Tamaño del subgrupo", "asis_n", 5, min_value=1, step=1)
    variable_n = checkbox_persistente(st, "El tamaño de muestra cambia", "asis_n_variable", False)
    specs = checkbox_persistente(st, "Tengo LIE y LSE", "asis_specs", False)

    recs = []
    if tipo == "Variable continua":
        if sub and 2 <= n <= 10:
            recs.append("Usa carta X-barra/R porque tienes subgrupos racionales de tamaño entre 2 y 10.")
        elif sub and n > 10:
            recs.append("Para subgrupos mayores a 10 se recomienda X-barra/S. Si solo tienes esta app, revisa I-MR o transforma el análisis por subgrupos.")
        else:
            recs.append("Usa carta I-MR porque estás trabajando con mediciones individuales o sin subgrupos racionales.")
        if specs:
            recs.append("Calcula capacidad: Cp, CPU, CPL, Cpk, Pp, Ppk, Cpm y producto no conforme.")
    elif tipo == "Defectuoso / no defectuoso":
        if variable_n:
            recs.append("Usa carta p porque cambia el tamaño de muestra inspeccionado.")
        else:
            recs.append("Usa carta np porque el tamaño de muestra es constante.")
    else:
        if variable_n:
            recs.append("Usa carta u porque cambia el número de unidades u oportunidades inspeccionadas.")
        else:
            recs.append("Usa carta c porque el área de oportunidad es constante.")
    recs.append("Valida normalidad para capacidad de variables. Valida independencia antes de interpretar cartas.")
    for i, r in enumerate(recs, 1):
        caja_estado(st, "ok", f"{i}. {r}")


def clasificar_indicador(valor, cortes, etiquetas):
    try:
        v = float(valor)
    except Exception:
        return etiquetas[-1]
    for limite, etiqueta in cortes:
        if v >= limite:
            return etiqueta
    return etiquetas[-1]


def generar_conclusion_experta(nombre_producto, variable, unidad, res_desc, normalidad_estado, independencia, capacidad, control_texto, objetivo_mejora, causas, acciones_extra):
    media = res_desc.get("Media", np.nan) if res_desc else np.nan
    sigma = res_desc.get("Desv. estándar", np.nan) if res_desc else np.nan
    n = res_desc.get("n", 0) if res_desc else 0
    partes = []
    partes.append(
        f"Se analizaron {int(n)} mediciones del producto o proceso {nombre_producto}, usando como variable crítica {variable}"
        + (f" en {unidad}" if unidad else "")
        + f". La media observada fue {fmt(media)} y la desviación estándar fue {fmt(sigma)}."
    )
    partes.append(f"El análisis de supuestos indica normalidad: {normalidad_estado}. La independencia temporal se clasifica como {independencia.get('Resultado', 'No evaluado')}.")
    partes.append(control_texto)
    if capacidad:
        partes.append(
            f"La capacidad del proceso se clasifica como {capacidad['Estado']}. Cp = {fmt(capacidad['Cp'])}, Cpk = {fmt(capacidad['Cpk'])} y el riesgo principal está hacia {capacidad['Riesgo principal']}. "
            f"El producto no conforme estimado total es {fmt(capacidad['% PNC estimado total'])}%, dividido en {fmt(capacidad['% PNC estimado LIE'])}% por LIE y {fmt(capacidad['% PNC estimado LSE'])}% por LSE."
        )
        if capacidad["Cpk"] < 1:
            partes.append("La conclusión técnica es que el proceso no debe liberarse como proceso capaz sin acciones correctivas, aunque sus cartas de control puedan mostrar estabilidad. Estar estable no significa cumplir especificaciones.")
        elif capacidad["Cpk"] < 1.33:
            partes.append("La conclusión técnica es que el proceso cumple de forma limitada. Debe mantenerse bajo vigilancia y mejorar su margen frente a las especificaciones.")
        else:
            partes.append("La conclusión técnica es favorable. El proceso tiene margen suficiente frente a las especificaciones, siempre que se mantenga estable en el tiempo.")
    else:
        partes.append("No se calcularon índices de capacidad porque no se ingresaron límites de especificación válidos. La conclusión se basa en estabilidad, tendencia, dispersión y comportamiento de la variable.")
    if causas:
        partes.append("Las causas probables que deben investigarse son: " + ", ".join(causas) + ".")
    if objetivo_mejora:
        partes.append("Objetivo de mejora definido: " + objetivo_mejora.strip())
    if acciones_extra:
        partes.append("Acciones específicas indicadas por el usuario: " + acciones_extra.strip())
    partes.append("La decisión recomendada es aplicar un plan de mejora, verificar el efecto con nuevos datos y repetir el análisis antes de cerrar el proyecto.")
    return "\n\n".join(partes)


def pantalla_conclusiones_mejora(st, go):
    df = st.session_state.get("df")
    if not dataframe_valido(df):
        st.warning("Primero carga datos.")
        return
    nums = columnas_numericas(df)
    if not nums:
        st.error("No hay columnas numéricas.")
        return

    st.markdown("### Conclusiones y plan de mejora")
    caja_estado(
        st,
        "info",
        "Este módulo resume el proyecto con criterio técnico. Sirve para procesos de alimentos, bebidas, manufactura, servicios, laboratorio, logística o cualquier variable de calidad."
    )

    guardar_estado_si_no_existe(st, "con_producto", "producto analizado")
    guardar_estado_si_no_existe(st, "con_unidad", "")
    guardar_estado_si_no_existe(st, "con_usar_specs", True)
    guardar_estado_si_no_existe(st, "con_cp_obj", 1.33)
    guardar_estado_si_no_existe(st, "con_pnc_obj", 5.0)
    guardar_estado_si_no_existe(st, "con_intervalo", 15.0)
    guardar_estado_si_no_existe(st, "con_objetivo", "Reducir la variabilidad, centrar el proceso y disminuir el producto no conforme.")
    guardar_estado_si_no_existe(st, "con_acciones_extra", "")

    c1, c2, c3 = st.columns([1.1, 1, 1])
    with c1:
        producto = st.text_input("Producto, servicio o proceso", key="con_producto")
        variable = selectbox_persistente(st, "Variable crítica", nums, "con_variable", nums[0])
    with c2:
        unidad = st.text_input("Unidad de medida", key="con_unidad", placeholder="ml, g, mm, kg, minutos...")
        usar_specs = checkbox_persistente(st, "Tengo límites de especificación", "con_usar_specs", True)
    with c3:
        objetivo_mejora = st.text_area("Objetivo de mejora", key="con_objetivo", height=86)

    s = convertir_a_numerica(obtener_columna(df, variable))
    if len(s) < 2:
        st.error("La variable seleccionada necesita al menos 2 datos numéricos.")
        return
    res_desc = resumen_descriptivo(s)
    normalidad = evaluar_normalidad(s)
    estado_norm, texto_norm = diagnostico_normalidad(normalidad)
    independencia = evaluar_independencia(s)

    lie = lse = vn = None
    capacidad = None
    col_specs, col_sigma, col_meta = st.columns(3)
    with col_specs:
        if usar_specs:
            guardar_estado_si_no_existe(st, "con_lie", float(s.min()))
            guardar_estado_si_no_existe(st, "con_lse", float(s.max()))
            guardar_estado_si_no_existe(st, "con_vn", float(s.mean()))
            vn = numero_persistente(st, "VN | objetivo", "con_vn", float(s.mean()))
            lie = numero_persistente(st, "LIE", "con_lie", float(s.min()))
            lse = numero_persistente(st, "LSE", "con_lse", float(s.max()))
    with col_sigma:
        metodo_sigma = radio_persistente(st, "Sigma para capacidad", ["Muestral", "Histórica"], "con_sigma_metodo", "Muestral", horizontal=True)
        sigma_hist = None
        if metodo_sigma == "Histórica":
            guardar_estado_si_no_existe(st, "con_sigma_hist", float(s.std(ddof=1)))
            sigma_hist = numero_persistente(st, "Sigma histórica", "con_sigma_hist", float(s.std(ddof=1)), min_value=0.000001, format="%.8f")
    with col_meta:
        cp_obj = numero_persistente(st, "Cpk objetivo", "con_cp_obj", 1.33, min_value=0.1, step=0.01)
        pnc_obj = numero_persistente(st, "% PNC máximo deseado", "con_pnc_obj", 5.0, min_value=0.0, max_value=100.0, step=0.1)

    if usar_specs and lie is not None and lse is not None and lie < lse:
        capacidad = calcular_capacidad(s, lie, lse, vn, sigma_hist, None)

    control_texto = "No se evaluó carta de control en este módulo. Usa Gráficos de control para confirmar estabilidad antes de cerrar la conclusión."
    subgrupos = [c for c in df.columns if c != variable]
    with st.expander("Evaluación opcional de estabilidad con subgrupos"):
        tipo_control = radio_persistente(st, "Tipo de evaluación", ["Automática simple", "I-MR", "X-barra/R"], "con_tipo_control", "Automática simple", horizontal=True)
        if tipo_control == "X-barra/R" and subgrupos:
            sub = selectbox_persistente(st, "Columna de subgrupo", subgrupos, "con_subgrupo", subgrupos[0])
            calc = calcular_xbar_r(df, variable, sub)
            if calc:
                _, lim = calc
                fuera_total = lim["Puntos fuera Xbarra"] + lim["Puntos fuera R"]
                control_texto = f"La revisión X-barra/R reporta {int(lim['Puntos fuera Xbarra'])} puntos fuera en medias y {int(lim['Puntos fuera R'])} puntos fuera en rangos."
                st.plotly_chart(grafico_xbar_r(go, df, variable, sub)[0], use_container_width=True, key="con_xbar")
                st.plotly_chart(grafico_xbar_r(go, df, variable, sub)[1], use_container_width=True, key="con_r")
            else:
                control_texto = "No se pudo calcular X-barra/R. Revisa que cada subgrupo tenga entre 2 y 10 datos."
        else:
            calc_imr = calcular_i_mr(s)
            if calc_imr:
                _, _, lim = calc_imr
                control_texto = f"La revisión I-MR reporta {int(lim['Puntos fuera I'])} puntos fuera en valores individuales y {int(lim['Puntos fuera MR'])} puntos fuera en rango móvil."
            else:
                control_texto = "No se pudo calcular I-MR por falta de datos o ausencia de variación."

    indicadores = {
        "n": res_desc["n"],
        "Media": res_desc["Media"],
        "Desv. estándar": res_desc["Desv. estándar"],
        "Normalidad": estado_norm,
        "Independencia": independencia["Resultado"],
    }
    if capacidad:
        indicadores.update({
            "Cp": capacidad["Cp"],
            "Cpk": capacidad["Cpk"],
            "% PNC total": capacidad["% PNC estimado total"],
            "Riesgo principal": capacidad["Riesgo principal"],
            "Estado": capacidad["Estado"],
        })
    tarjetas(st, indicadores, "Resumen ejecutivo")

    causas = []
    if capacidad:
        if capacidad["Cpk"] < cp_obj:
            causas.append("capacidad insuficiente frente al objetivo")
        if capacidad["Centrado"] == "Descentrado":
            causas.append("media descentrada respecto al objetivo")
        if capacidad["% PNC estimado total"] > pnc_obj:
            causas.append("porcentaje de no conformes mayor al permitido")
        if capacidad["Riesgo principal"] == "LSE":
            causas.append("riesgo de valores altos o sobrellenado")
        elif capacidad["Riesgo principal"] == "LIE":
            causas.append("riesgo de valores bajos o subllenado")
    if independencia.get("Resultado") == "No cumple":
        causas.append("posible dependencia temporal, tendencia o ciclos")
    if estado_norm == "No cumple":
        causas.append("distribución no normal o presencia de mezcla de condiciones")
    if not causas:
        causas.append("no se observan señales críticas con los criterios seleccionados")

    acciones_base = []
    if capacidad:
        if capacidad["Cpk"] < 1:
            acciones_base.append("No liberar el proceso como capaz hasta aplicar correcciones y verificar nuevos datos.")
        if capacidad["Centrado"] == "Descentrado":
            acciones_base.append("Ajustar el set point o parámetro central del proceso hacia el VN.")
        if capacidad["Cp"] < cp_obj:
            acciones_base.append("Reducir variabilidad con calibración, mantenimiento, estandarización del método y control de materia prima.")
        if capacidad["Riesgo principal"] == "LSE":
            acciones_base.append("Controlar el exceso hacia el límite superior. Revisar dosificación, presión, velocidad, desgaste, temperatura o ajuste operativo según el tipo de producto.")
        elif capacidad["Riesgo principal"] == "LIE":
            acciones_base.append("Controlar el defecto hacia el límite inferior. Revisar suministro, dosificación, falta de material, tiempo de ciclo o ajuste operativo.")
    acciones_base.append("Recolectar una nueva corrida después de los ajustes y repetir cartas de control, capacidad y PNC.")
    acciones_base.append("Definir responsable, fecha de revisión, criterio de aceptación y evidencia antes de cerrar el proyecto.")

    acciones_extra = st.text_area("Acciones propias del caso", key="con_acciones_extra", height=80, placeholder="Ejemplo: calibrar dosificadores, cambiar boquillas, capacitar operarios...")

    tab1, tab2, tab3, tab4 = st.tabs(["Diagnóstico", "Plan de mejora", "Conclusión lista", "Gráficos"])
    with tab1:
        st.markdown("#### Hallazgos principales")
        for causa in causas:
            caja_estado(st, "alerta" if "no se observan" not in causa else "ok", causa)
        st.markdown("#### Supuestos")
        st.dataframe(normalidad, use_container_width=True, hide_index=True)
        tarjetas(st, independencia, "Independencia")
    with tab2:
        plan = pd.DataFrame({
            "Prioridad": list(range(1, len(acciones_base) + 1)),
            "Acción recomendada": acciones_base,
            "Propósito": [
                "Evitar decisiones con proceso no capaz" if i == 0 else
                "Mejorar cumplimiento y reducir no conformes" if "Ajustar" in acc or "Reducir" in acc else
                "Atacar el lado de mayor riesgo" if "límite" in acc else
                "Confirmar que la mejora funcionó" if "nueva corrida" in acc else
                "Cerrar el ciclo de mejora con evidencia"
                for i, acc in enumerate(acciones_base)
            ],
        })
        st.dataframe(plan, use_container_width=True, hide_index=True)
    with tab3:
        conclusion = generar_conclusion_experta(producto, variable, unidad, res_desc, estado_norm, independencia, capacidad, control_texto, objetivo_mejora, causas, acciones_extra)
        st.text_area("Conclusión para el informe", conclusion, height=360)
        caja_estado(st, "info", "Copia este texto y ajústalo con nombres de máquinas, turnos, clientes o normas cuando aplique.")
    with tab4:
        st.plotly_chart(grafico_corrida(go, s, "Comportamiento de la variable crítica"), use_container_width=True, key="con_corrida")
        if usar_specs and lie is not None and lse is not None and lie < lse:
            st.plotly_chart(grafico_hist_capacidad(go, s, lie, lse, vn), use_container_width=True, key="con_hist")

def escribir_titulo_informe(ws, fila, titulo):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=8)
    c = ws.cell(fila, 1)
    c.value = titulo
    c.font = Font(bold=True, color="FFFFFF", size=13)
    c.fill = PatternFill("solid", fgColor="1D4ED8")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[fila].height = 24
    return fila + 2


def escribir_tabla_informe(ws, fila, titulo, df):
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    fila = escribir_titulo_informe(ws, fila, titulo)
    borde = Border(left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"), top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="CBD5E1"))
    for col_idx, nombre in enumerate(df.columns, start=1):
        celda = ws.cell(fila, col_idx)
        celda.value = nombre
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="0F172A")
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.border = borde
    for r_idx, row in enumerate(df.itertuples(index=False), start=fila + 1):
        for c_idx, valor in enumerate(row, start=1):
            celda = ws.cell(r_idx, c_idx)
            if isinstance(valor, (float, int, np.integer, np.floating)) and not pd.isna(valor):
                celda.value = valor_limpio(valor)
                celda.number_format = "General"
            else:
                celda.value = valor
            celda.border = borde
            celda.alignment = Alignment(vertical="center", wrap_text=True)
    return fila + len(df) + 3


def ajustar_informe_excel(ws):
    from openpyxl.utils import get_column_letter
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    anchos = [26, 20, 20, 20, 20, 20, 20, 20]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


def insertar_imagen_figura(ws, fig, celda, texto_fallo):
    try:
        from openpyxl.drawing.image import Image as XLImage
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
        tmp.close()
        fig.write_image(tmp.name, width=980, height=580, scale=1)
        img = XLImage(tmp.name)
        img.width = 680
        img.height = 400
        ws.add_image(img, celda)
        return True
    except Exception:
        ws[celda] = texto_fallo
        return False


def pantalla_reporte(st, go):
    df = st.session_state.get("df")
    if not dataframe_valido(df):
        st.warning("Primero carga datos.")
        return

    nums = columnas_numericas(df)
    if not nums:
        st.error("No hay columnas numéricas.")
        return

    st.markdown("### Informe visual en Excel")
    caja_estado(st, "info", "El informe se genera en una sola hoja. Usa los mismos gráficos de la app como imágenes. Para insertar imágenes instala: python -m pip install kaleido")

    c1, c2, c3 = st.columns(3)
    with c1:
        variable = selectbox_persistente(st, "Variable principal", nums, "rep_variable", nums[0])
        subgrupo = selectbox_persistente(st, "Subgrupo para carta X-barra/R", ["Ninguno"] + df.columns.tolist(), "rep_subgrupo", "Ninguno")
    with c2:
        guardar_estado_si_no_existe(st, "rep_vn", float(convertir_a_numerica(obtener_columna(df, variable)).mean()))
        vn = numero_persistente(st, "VN | Valor nominal", "rep_vn", float(convertir_a_numerica(obtener_columna(df, variable)).mean()))
        guardar_estado_si_no_existe(st, "rep_lie", float(convertir_a_numerica(obtener_columna(df, variable)).min()))
        lie = numero_persistente(st, "LIE | Límite inferior", "rep_lie", float(convertir_a_numerica(obtener_columna(df, variable)).min()))
    with c3:
        guardar_estado_si_no_existe(st, "rep_lse", float(convertir_a_numerica(obtener_columna(df, variable)).max()))
        lse = numero_persistente(st, "LSE | Límite superior", "rep_lse", float(convertir_a_numerica(obtener_columna(df, variable)).max()))
        guardar_estado_si_no_existe(st, "rep_sigma_hist", 0.0)
        sigma_hist = numero_persistente(st, "Sigma histórica opcional", "rep_sigma_hist", 0.0, min_value=0.0, format="%.6f")

    s = convertir_a_numerica(obtener_columna(df, variable))
    sigma_usada = sigma_hist if sigma_hist > 0 else None
    normalidad = evaluar_normalidad(s)
    estado_norm, texto_norm = diagnostico_normalidad(normalidad)
    resumen = pd.DataFrame([resumen_descriptivo(s)]).T.reset_index()
    resumen.columns = ["Indicador", "Valor"]
    cap = calcular_capacidad(s, lie, lse, vn, sigma_usada, None) if lie < lse else None

    if cap is not None:
        capacidad_df = pd.DataFrame([redondear_dict(cap)]).T.reset_index()
        capacidad_df.columns = ["Indicador", "Valor"]
        pnc_df = pd.DataFrame({
            "Indicador": ["% PNC estimado LIE", "% PNC estimado LSE", "% PNC estimado total", "PPM estimado", "PNC observado"],
            "Valor": [cap["% PNC estimado LIE"], cap["% PNC estimado LSE"], cap["% PNC estimado total"], cap["PPM estimado"], cap["Producto no conforme observado"]],
        })
    else:
        capacidad_df = pd.DataFrame({"Indicador": ["Capacidad"], "Valor": ["No calculada. Revisa LIE y LSE."]})
        pnc_df = pd.DataFrame({"Indicador": ["PNC"], "Valor": ["No calculado"]})

    config_df = pd.DataFrame({
        "Campo": ["Aplicación", "Variable", "Subgrupo", "VN", "LIE", "LSE", "Normalidad", "Fecha"],
        "Valor": ["Pulso de Calidad SPC", variable, subgrupo, vn, lie, lse, estado_norm, pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")],
    })

    salida = BytesIO()
    with pd.ExcelWriter(salida, engine="openpyxl") as writer:
        pd.DataFrame({" ": []}).to_excel(writer, index=False, sheet_name="Informe")
        wb = writer.book
        ws = wb["Informe"]
        ws.delete_rows(1, 2)

        fila = 1
        fila = escribir_titulo_informe(ws, fila, "PULSO DE CALIDAD SPC - INFORME DE CONTROL ESTADÍSTICO")
        fila = escribir_tabla_informe(ws, fila, "Configuración del análisis", config_df)
        fila = escribir_tabla_informe(ws, fila, "Resumen estadístico", resumen)
        fila = escribir_tabla_informe(ws, fila, "Supuesto de normalidad", normalidad)
        ws.cell(fila, 1).value = "Diagnóstico de normalidad"
        ws.cell(fila, 2).value = texto_norm
        fila += 3
        fila = escribir_tabla_informe(ws, fila, "Análisis de capacidad", capacidad_df)
        fila = escribir_tabla_informe(ws, fila, "Producto no conforme", pnc_df)

        if subgrupo != "Ninguno":
            calc = calcular_xbar_r(df, variable, subgrupo)
            if calc is not None:
                resumen_x, lim_x = calc
                tabla_x = resumen_x.reset_index().rename(columns={"mean": "Xbarra"})
                fila = escribir_tabla_informe(ws, fila, "Datos de carta X-barra y R", tabla_x.head(35))
                fig_control, fig_r, _ = grafico_xbar_r(go, df, variable, subgrupo)
            else:
                fig_control = grafico_corrida(go, s, "Gráfico de corrida")
                fig_r = None
        else:
            figs = grafico_i_mr(go, s)
            fig_control = figs[0] if figs[0] is not None else grafico_corrida(go, s, "Gráfico de corrida")
            fig_r = figs[1] if figs[1] is not None else None

        fila = escribir_titulo_informe(ws, fila, "Gráficos generados desde la app")
        insertar_imagen_figura(ws, grafico_qq(go, s), f"A{fila}", "No se pudo insertar el gráfico de normalidad. Instala kaleido.")
        insertar_imagen_figura(ws, grafico_hist_capacidad(go, s, lie, lse, vn), f"A{fila + 23}", "No se pudo insertar el gráfico de capacidad. Instala kaleido.")
        insertar_imagen_figura(ws, fig_control, f"A{fila + 46}", "No se pudo insertar el gráfico de control. Instala kaleido.")
        if fig_r is not None:
            insertar_imagen_figura(ws, fig_r, f"A{fila + 69}", "No se pudo insertar la segunda carta de control. Instala kaleido.")
        ajustar_informe_excel(ws)

    tarjetas(st, {"Variable": variable, "Normalidad": estado_norm, "LIE": lie, "LSE": lse}, "Contenido del informe")
    st.download_button(
        "Descargar informe Excel",
        data=salida.getvalue(),
        file_name="informe_pulso_calidad_spc.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ============================================================
# PRUEBAS
# ============================================================

def ejecutar_pruebas():
    if FALTAN_BASE:
        print(mensaje_dependencias(FALTAN_BASE))
        return False
    s = pd.Series([10, 10.2, 9.8, 10.1, 9.9, 10.3, 10.0, 9.7])
    assert resumen_descriptivo(s)["n"] == 8
    assert evaluar_normalidad(s).shape[0] >= 1
    assert "Durbin-Watson" in evaluar_independencia(s)
    assert calcular_capacidad(s, 9, 11, 10) is not None
    df = pd.DataFrame({"subgrupo": [1, 1, 2, 2], "medicion": [10, 11, 9, 10]})
    assert calcular_xbar_r(df, "medicion", "subgrupo") is not None
    assert calcular_i_mr(s) is not None
    print("Todas las pruebas internas pasaron correctamente.")
    return True


# ============================================================
# APP
# ============================================================

def ejecutar_app():
    if FALTAN_BASE:
        print(mensaje_dependencias(FALTAN_BASE))
        return
    st, go = importar_librerias_app()
    if st is None or go is None:
        print(mensaje_dependencias(FALTAN_APP))
        return
    st.set_page_config(page_title="Pulso de Calidad SPC", page_icon="📊", layout="wide")
    aplicar_estilo_visual(st)
    if "df" not in st.session_state:
        st.session_state["df"] = None
    if "df_original" not in st.session_state:
        st.session_state["df_original"] = None
    st.sidebar.title("Pulso de Calidad SPC")
    st.sidebar.caption("Control Estadístico de Procesos")
    menu = st.sidebar.radio("Selecciona un módulo", [
        "Inicio",
        "Cargar datos",
        "Supuestos del proceso",
        "Gráficos de control",
        "Capacidad del proceso",
        "Producto no conforme",
        "Muestreo de aceptación",
        "Asistente de proyecto",
        "Conclusiones y plan de mejora",
        "Reporte",
    ], key="menu_principal")
    encabezado(st, "Pulso de Calidad SPC", "Panel visual para analizar, controlar y mejorar procesos con datos.")
    if menu == "Inicio":
        pantalla_inicio(st)
    elif menu == "Cargar datos":
        pantalla_cargar_datos(st)
    elif menu == "Supuestos del proceso":
        pantalla_supuestos(st, go)
    elif menu == "Gráficos de control":
        pantalla_control(st, go)
    elif menu == "Capacidad del proceso":
        pantalla_capacidad(st, go)
    elif menu == "Producto no conforme":
        pantalla_no_conformes(st, go)
    elif menu == "Muestreo de aceptación":
        pantalla_muestreo(st, go)
    elif menu == "Asistente de proyecto":
        pantalla_asistente(st)
    elif menu == "Conclusiones y plan de mejora":
        pantalla_conclusiones_mejora(st, go)
    elif menu == "Reporte":
        pantalla_reporte(st, go)


if __name__ == "__main__":
    if "--self-test" in sys.argv:
        ejecutar_pruebas()
    else:
        ejecutar_app()
